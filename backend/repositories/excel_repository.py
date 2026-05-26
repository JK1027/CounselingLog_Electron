import os
import shutil
import datetime
import uuid
import tempfile
import threading
import time
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from backend.utils.path_helper import get_resource_path, get_writable_path, ensure_directory_exists, get_user_backup_path
from backend.utils.security import sanitize_value
from backend.utils.logger import logger
from backend.core.constants import (
    SHEET_NAMES, 
    GROUP_COUNSELING_SHEET, 
    REQUEST_SHEET, 
    RECORD_ID_COL, 
    SHEET_SCHEMAS
)
from backend.utils.excel_helpers import (
    find_real_max_row, 
    find_empty_row_by_key, 
    apply_excel_formatting,
    clone_cell_style,
    update_data_validation_ranges,
    update_auto_filter_range
)

class ExcelRepository:
    """엑셀 파일 데이터를 로드, 수정, 저장 및 백업을 담당하는 리포지토리 클래스"""
    def __init__(self, main_file_path=None):
        self.main_file_path = main_file_path
        self.data_frames = {sheet: pd.DataFrame() for sheet in SHEET_NAMES}
        self.lock = threading.RLock()
        self._last_loaded_time = 0.0
        self._last_checked_time = 0.0
        self.throttle_interval = 2.0
        self.dirty_uuid = False
        
        # Validation options loading from CONFIG
        from backend.core.constants import CONFIG
        self.validation_options = CONFIG.get("validation_options", {})
        self.validation_options_cache = {}

    def sanitize_value(self, value):
        """Excel Formula Injection 방어를 위해 security 모듈의 sanitize_value를 호출합니다."""
        return sanitize_value(value)

    def load_data(self, file_path):
        """지정된 경로의 엑셀 파일을 로드하고, UUID가 누락된 경우 자동 생성하여 저장합니다."""
        with self.lock:
            self.main_file_path = file_path
            uuid_updated = False

            with pd.ExcelFile(file_path) as xls:
                excel_sheets_in_file = xls.sheet_names
                
                for sheet_name in SHEET_NAMES:
                    if sheet_name in excel_sheets_in_file:
                        df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str).fillna('')
                        if '학번' in df.columns:
                            df['학번'] = df['학번'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                        
                        # UUID 컬럼이 존재하지 않는 경우 추가
                        if RECORD_ID_COL not in df.columns:
                            df[RECORD_ID_COL] = ""
                            
                        # 비어있는 UUID 값을 채움
                        for idx, row in df.iterrows():
                            if not str(row[RECORD_ID_COL]).strip():
                                df.at[idx, RECORD_ID_COL] = uuid.uuid4().hex
                                uuid_updated = True
                                
                        self.data_frames[sheet_name] = df
                    else:
                        # 새로운 시트의 빈 DataFrame 생성 시 기본 스키마 반영
                        schema = SHEET_SCHEMAS.get(sheet_name, [])
                        self.data_frames[sheet_name] = pd.DataFrame(columns=schema)

            if os.path.exists(file_path):
                self._last_loaded_time = os.path.getmtime(file_path)
                self._last_checked_time = time.time()

            logger.info(f"데이터 로드 완료: {file_path}")
            
            # 유효성 검사 드롭다운 옵션 파싱 및 캐시 갱신
            from backend.utils.validation_parser import parse_validation_options
            self.validation_options_cache = parse_validation_options(file_path, self.validation_options)
            
            # 만약 신규 생성된 UUID가 있다면 dirty_uuid 플래그 마킹 (Startup 지연 방지)
            if uuid_updated:
                self.dirty_uuid = True
                logger.info("누락된 UUID 식별키를 발견하여 자동 보정했습니다. (디스크 저장 지연 적용)")
                
            return True

    def check_and_reload(self):
        """파일이 외부에서 수정되었거나 최초 로드 시에만 메모리 캐시를 리로드합니다."""
        if not self.main_file_path or not os.path.exists(self.main_file_path):
            return
        with self.lock:
            now = time.time()
            if now - self._last_checked_time < self.throttle_interval:
                return  # 스로틀링: 2초 내에는 디스크 I/O 생략
            
            try:
                current_mtime = os.path.getmtime(self.main_file_path)
                self._last_checked_time = now
                if current_mtime != self._last_loaded_time:
                    logger.info("Excel 파일 외부 변경 감지: 메모리 캐시를 리로드합니다.")
                    self.load_data(self.main_file_path)
            except Exception as e:
                logger.error(f"check_and_reload 도중 예외 발생: {e}")
                # Fallback: 이미 메모리에 정상적으로 데이터가 로드된 상태라면 에러를 상위로 던지지 않고 서빙을 유지함.
                if any(not df.empty for df in self.data_frames.values()):
                    logger.warning("리로드 오류로 인해 이전 메모리 캐시 데이터를 유지하여 서빙합니다.")
                else:
                    raise e

    def get_validation_options(self):
        """캐싱된 데이터 유효성 검사 드롭박스 옵션을 반환합니다."""
        with self.lock:
            if not self.validation_options_cache:
                from backend.utils.validation_parser import parse_validation_options
                self.validation_options_cache = parse_validation_options(self.main_file_path, self.validation_options)
            return self.validation_options_cache


    def swap_temp_and_original(self, temp_file_path, file_path):
        """
        임시 파일을 원본 파일로 교체합니다. 교체 전 원본 파일은 backup 폴더에 복사해 둡니다 (Safe Save).
        """
        try:
            if not os.path.exists(temp_file_path) or os.path.getsize(temp_file_path) == 0:
                raise IOError("임시 파일 저장 확인 실패 (파일이 없거나 비어 있음)")
            
            # 덮어쓰기 전 안전을 위해 자동 백업본 생성
            if os.path.exists(file_path):
                backup_dir = get_writable_path("backup")
                ensure_directory_exists(backup_dir)
                backup_time = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
                name, ext = os.path.splitext(os.path.basename(file_path))
                backup_path = os.path.join(backup_dir, f"{name}_auto_backup_{backup_time}{ext}")
                shutil.copy2(file_path, backup_path)
                logger.info(f"Safe Save: 원본 파일 백업 생성 완료 -> {backup_path}")
            
            # 원자적 파일 교체 수행
            try:
                os.replace(temp_file_path, file_path)
            except PermissionError:
                raise PermissionError(f"'{os.path.basename(file_path)}' 파일이 다른 프로그램에서 열려있어 덮어쓸 수 없습니다.\n파일을 닫고 다시 시도해주세요.")
            
            return True, None
        except Exception as e:
            logger.error(f"임시 파일 스왑 실패: {e}")
            if os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except Exception:
                    pass
            return False, str(e)

    def restore_latest_backup(self):
        """저장 오류 시 최근의 자동 백업본을 찾아 복원(Rollback)을 수행합니다."""
        logger.warning("오류 감지: 최근 백업 데이터로부터 자동 복원(Rollback)을 개시합니다.")
        try:
            backup_dir = get_writable_path("backup")
            if not os.path.exists(backup_dir):
                logger.error("복원 실패: 백업 폴더가 존재하지 않습니다.")
                return False
                
            files = [f for f in os.listdir(backup_dir) if "auto_backup" in f and f.endswith(".xlsx")]
            if not files:
                logger.error("복원 실패: 복구 가능한 자동 백업 파일을 찾을 수 없습니다.")
                return False
                
            # 가장 생성 날짜가 최신인 파일 찾기
            files.sort(reverse=True)
            latest_backup = os.path.join(backup_dir, files[0])
            
            shutil.copy2(latest_backup, self.main_file_path)
            logger.info(f"성공적으로 백업본 복구 완료: {latest_backup} -> {self.main_file_path}")
            
            # 메모리도 해당 백업 기준으로 리로드
            self.load_data(self.main_file_path)
            return True
        except Exception as ex:
            logger.critical(f"자동 복원 롤백 도중 심각한 예외 발생: {ex}")
            return False

    def _find_matching_excel_row(self, worksheet, sheet_name, original_data):
        """
        수정 전 메모리 행 데이터(original_data)와 일치하는 행 번호를 3단계 Fallback 구조로 추적합니다.
        """
        headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
        real_max = find_real_max_row(worksheet)
        
        # --- 1순위: UUID (_record_id) 매칭 ---
        if RECORD_ID_COL in headers:
            col_idx = headers[RECORD_ID_COL]
            target_uuid = str(original_data.get(RECORD_ID_COL, "")).strip()
            if target_uuid:
                matched_rows = []
                for r in range(2, real_max + 1):
                    excel_uuid = str(worksheet.cell(row=r, column=col_idx).value or "").strip()
                    if excel_uuid == target_uuid:
                        matched_rows.append(r)
                if len(matched_rows) == 1:
                    return matched_rows[0]

        # --- 2순위 Fallback: (학번/이름) + 상담일자 + 상담회기 매칭 ---
        key_field = "학번" if sheet_name == GROUP_COUNSELING_SHEET else "이름"
        if key_field in headers and "*상담일자" in headers and "상담회기" in headers:
            key_idx = headers[key_field]
            date_idx = headers["*상담일자"]
            session_idx = headers["상담회기"]
            
            target_key = str(original_data.get(key_field, "")).strip().replace(".0", "")
            target_date = str(original_data.get("*상담일자", "")).strip()
            target_session = str(original_data.get("상담회기", "")).strip()
            
            matched_rows = []
            for r in range(2, real_max + 1):
                excel_key = str(worksheet.cell(row=r, column=key_idx).value or "").strip().replace(".0", "")
                excel_date = str(worksheet.cell(row=r, column=date_idx).value or "").strip()
                excel_session = str(worksheet.cell(row=r, column=session_idx).value or "").strip()
                
                if excel_key == target_key and excel_date == target_date and excel_session == target_session:
                    matched_rows.append(r)
            if len(matched_rows) == 1:
                logger.info(f"Fallback 매칭 작동 (2순위): 행 번호 {matched_rows[0]} 매칭 완료")
                return matched_rows[0]

        # --- 3순위 Fallback: 기존 전체 필드 비교 방식 ---
        matched_rows = []
        for r in range(2, real_max + 1):
            is_match = True
            for col_name, col_idx in headers.items():
                if col_name not in original_data or col_name == RECORD_ID_COL:
                    continue
                excel_val = worksheet.cell(row=r, column=col_idx).value
                orig_val = str(original_data[col_name]).strip()
                excel_val_str = str(excel_val).strip() if excel_val is not None else ""
                
                if col_name == '학번':
                    orig_val = orig_val.replace(".0", "")
                    excel_val_str = excel_val_str.replace(".0", "")
                
                if orig_val != excel_val_str:
                    is_match = False
                    break
            
            if is_match:
                matched_rows.append(r)
                
        if not matched_rows:
            raise ValueError("수정할 대상 행을 엑셀 파일에서 찾을 수 없습니다. (데이터 불일치)")
        if len(matched_rows) > 1:
            raise ValueError(f"동일한 내용의 데이터가 여러 개 발견되어 대상을 특정할 수 없습니다. (매칭된 행: {matched_rows})")
            
        logger.info(f"Fallback 매칭 작동 (3순위): 행 번호 {matched_rows[0]} 매칭 완료")
        return matched_rows[0]

    def save_all_data_to_excel(self):
        """메모리에 있는 모든 데이터프레임을 기존 엑셀 파일 서식(드롭다운, 필터 등)을 보존하며 안전하게 덮어씁니다 (Safe Save & NamedTemporaryFile 적용)."""
        with self.lock:
            if not self.main_file_path:
                return False, "저장할 파일 경로가 지정되지 않았습니다."
            
            ext = os.path.splitext(self.main_file_path)[1]
            target_dir = os.path.dirname(self.main_file_path) or "."
            
            # 동일 드라이브/볼륨 상에 안전한 임시 파일 생성
            temp_file = tempfile.NamedTemporaryFile(dir=target_dir, delete=False, suffix=ext)
            temp_file_path = temp_file.name
            temp_file.close() # openpyxl에서 쓰기 위해 일단 닫음
            
            workbook = None
            try:
                # 기존 엑셀 파일 서식을 보존하기 위해 복사, 없을 시 기본 템플릿 생성 후 복사
                if os.path.exists(self.main_file_path):
                    shutil.copy2(self.main_file_path, temp_file_path)
                else:
                    self._ensure_template_initialized()
                    shutil.copy2(self.main_file_path, temp_file_path)
                
                workbook = openpyxl.load_workbook(temp_file_path)
                
                for sheet_name, df in self.data_frames.items():
                    if sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                    else:
                        worksheet = workbook.create_sheet(title=sheet_name)
                    
                    # 1. 헤더 동기화 및 맵핑
                    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
                    if not headers:
                        for col_idx, col_name in enumerate(df.columns, 1):
                            worksheet.cell(row=1, column=col_idx, value=col_name)
                        headers = {col_name: col_idx for col_idx, col_name in enumerate(df.columns, 1)}
                    else:
                        for col_name in df.columns:
                            if col_name not in headers:
                                new_col_idx = worksheet.max_column + 1
                                worksheet.cell(row=1, column=new_col_idx, value=col_name)
                                headers[col_name] = new_col_idx
                    
                    real_max = find_real_max_row(worksheet)
                    new_max = len(df) + 1
                    
                    # 2. 데이터 업데이트 및 새 행 스타일 복제
                    for row_idx, row_series in df.iterrows():
                        excel_row = row_idx + 2
                        is_new_row = (excel_row > real_max)
                        
                        if is_new_row:
                            # 윗 행(또는 데이터 시작 행인 2행)을 템플릿으로 참조하여 스타일/높이 복제
                            ref_row = excel_row - 1 if excel_row - 1 >= 2 else 2
                            if worksheet.row_dimensions[ref_row].height is not None:
                                worksheet.row_dimensions[excel_row].height = worksheet.row_dimensions[ref_row].height
                        
                        for col_name, value in row_series.items():
                            col_idx = headers[col_name]
                            cell = worksheet.cell(row=excel_row, column=col_idx)
                            
                            # 새 행인 경우에만 스타일 복제 수행
                            if is_new_row:
                                ref_cell = worksheet.cell(row=ref_row, column=col_idx)
                                clone_cell_style(ref_cell, cell)
                                
                            cell.value = self.sanitize_value(value)
                    
                    # 3. 유효성 검사 및 자동 필터 범위 자동 재조정
                    update_data_validation_ranges(worksheet, real_max, new_max)
                    update_auto_filter_range(worksheet, new_max)
                    
                    # 4. 남는 잔여 행만 값 지우기 (성능 최적화)
                    if real_max > new_max:
                        for r in range(new_max + 1, real_max + 1):
                            worksheet.row_dimensions[r].height = None
                            for c in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=r, column=c).value = None
                    
                    # 5. _record_id 컬럼 숨김 처리
                    for cell in worksheet[1]:
                        if cell.value == RECORD_ID_COL:
                            from openpyxl.utils import get_column_letter
                            worksheet.column_dimensions[get_column_letter(cell.column)].hidden = True
                            break
                
                workbook.save(temp_file_path)
                workbook.close()
                workbook = None
                
                success, err = self.swap_temp_and_original(temp_file_path, self.main_file_path)
                if success:
                    if os.path.exists(self.main_file_path):
                        self._last_loaded_time = os.path.getmtime(self.main_file_path)
                        self._last_checked_time = time.time()
                    self.dirty_uuid = False
                    logger.info("모든 데이터프레임이 기존 서식을 보존하며 안전하게 덮어쓰기 완료되었습니다.")
                    return True, None
                return False, err
            except PermissionError:
                return False, f"'{os.path.basename(self.main_file_path)}' 파일이 다른 프로그램에서 열려있어 저장할 수 없습니다.\n파일을 닫고 다시 시도해주세요."
            except Exception as e:
                logger.error(f"save_all_data_to_excel 에러: {e}")
                return False, str(e)
            finally:
                if workbook is not None:
                    try:
                        workbook.close()
                    except Exception:
                        pass
                if os.path.exists(temp_file_path):
                    try:
                        os.remove(temp_file_path)
                    except Exception:
                        pass

    def update_excel_row(self, sheet_name, df_index, updated_data):
        """지정된 행의 데이터를 openpyxl을 사용하여 엑셀 파일에서 직접 수정하고, 메모리 데이터프레임도 즉시 동기화합니다."""
        with self.lock:
            if not self.main_file_path:
                return False, "엑셀 파일 경로를 찾을 수 없습니다."
            
            df = self.data_frames.get(sheet_name)
            if df is None or df_index >= len(df):
                return False, f"메모리 데이터프레임에서 인덱스 {df_index}를 찾을 수 없습니다."
            original_data = df.iloc[df_index].to_dict()
            
            ext = os.path.splitext(self.main_file_path)[1]
            target_dir = os.path.dirname(self.main_file_path) or "."
            temp_file = tempfile.NamedTemporaryFile(dir=target_dir, delete=False, suffix=ext)
            temp_file_path = temp_file.name
            temp_file.close()

            workbook = None
            try:
                shutil.copy2(self.main_file_path, temp_file_path)
                workbook = openpyxl.load_workbook(temp_file_path)
                worksheet = workbook[sheet_name]
                
                try:
                    excel_row_num = self._find_matching_excel_row(worksheet, sheet_name, original_data)
                except ValueError as ve:
                    return False, str(ve)
                
                headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
                
                sanitized_updates = {}
                for col_name, value in updated_data.items():
                    if col_name in headers:
                        col_idx = headers[col_name]
                        sanitized_val = self.sanitize_value(value)
                        worksheet.cell(row=excel_row_num, column=col_idx).value = sanitized_val
                        sanitized_updates[col_name] = sanitized_val
                
                # 상담회기 재정렬
                target_name = str(original_data.get("이름", "")).strip()
                target_sid = str(original_data.get("학번", "")).strip()
                self._resequence_sessions_in_worksheet(worksheet, sheet_name, target_name, target_sid)

                workbook.save(temp_file_path)
                workbook.close()
                workbook = None
    
                success, err = self.swap_temp_and_original(temp_file_path, self.main_file_path)
                if not success:
                    # 덮어쓰기 등 스왑 도중 실패 시 자동 롤백 적용
                    self.restore_latest_backup()
                    return False, err

                # Sync in-memory DataFrame
                for col, val in sanitized_updates.items():
                    df.at[df_index, col] = val

                # 메모리 데이터프레임 상담회기 재정렬
                df = self._resequence_sessions_in_df(df, sheet_name, target_name, target_sid)
                self.data_frames[sheet_name] = df
                
                if os.path.exists(self.main_file_path):
                    self._last_loaded_time = os.path.getmtime(self.main_file_path)
                    self._last_checked_time = time.time()
                    
                logger.info(f"Memory Sync 완료: [{sheet_name}] row={excel_row_num} (df_index={df_index}) 수정")
                return True, None

            except PermissionError:
                return False, f"'{os.path.basename(self.main_file_path)}' 파일이 다른 프로그램에서 열려있어 수정할 수 없습니다.\n파일을 닫고 다시 시도해주세요."
            except Exception as e:
                logger.error(f"update_excel_row 에러: {e}")
                self.restore_latest_backup()  # 장애 발생 시 백업 롤백
                return False, str(e)
            finally:
                if workbook is not None:
                    try:
                        workbook.close()
                    except Exception:
                        pass
                if os.path.exists(temp_file_path):
                    try:
                        os.remove(temp_file_path)
                    except Exception:
                        pass

    def delete_excel_row(self, sheet_name, df_index):
        """지정된 행을 엑셀 파일에서 삭제하고 메모리 데이터프레임에서도 동기화합니다. 순번이 존재하는 경우 순번을 재정렬합니다."""
        with self.lock:
            if not self.main_file_path:
                return False, "엑셀 파일 경로를 찾을 수 없습니다."
            
            df = self.data_frames.get(sheet_name)
            if df is None or df_index >= len(df):
                return False, f"메모리 데이터프레임에서 인덱스 {df_index}를 찾을 수 없습니다."
                
            original_data = df.iloc[df_index].to_dict()
            
            ext = os.path.splitext(self.main_file_path)[1]
            target_dir = os.path.dirname(self.main_file_path) or "."
            temp_file = tempfile.NamedTemporaryFile(dir=target_dir, delete=False, suffix=ext)
            temp_file_path = temp_file.name
            temp_file.close()

            workbook = None
            try:
                shutil.copy2(self.main_file_path, temp_file_path)
                workbook = openpyxl.load_workbook(temp_file_path)
                worksheet = workbook[sheet_name]
                
                try:
                    excel_row_num = self._find_matching_excel_row(worksheet, sheet_name, original_data)
                except ValueError as ve:
                    return False, str(ve)
                
                # 행 삭제
                worksheet.delete_rows(excel_row_num, 1)
                
                # 순번 재정렬
                headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
                if "순번" in headers:
                    seq_col_idx = headers["순번"]
                    key_col_name = "학번" if sheet_name == GROUP_COUNSELING_SHEET else "이름"
                    if key_col_name in headers:
                        key_col_idx = headers[key_col_name]
                        seq_counter = 1
                        real_max = find_real_max_row(worksheet)
                        for r in range(2, real_max + 1):
                            seq_cell = worksheet.cell(row=r, column=seq_col_idx)
                            key_val = worksheet.cell(row=r, column=key_col_idx).value
                            
                            if str(seq_cell.value).strip() == "예시":
                                continue
                                
                            if key_val is not None and str(key_val).strip() != "":
                                seq_cell.value = str(seq_counter)
                                seq_counter += 1

                # 상담회기 재정렬
                target_name = str(original_data.get("이름", "")).strip()
                target_sid = str(original_data.get("학번", "")).strip()
                self._resequence_sessions_in_worksheet(worksheet, sheet_name, target_name, target_sid)

                workbook.save(temp_file_path)
                workbook.close()
                workbook = None
     
                success, err = self.swap_temp_and_original(temp_file_path, self.main_file_path)
                if not success:
                    self.restore_latest_backup()
                    return False, err

                # 메모리 데이터프레임 동기화
                df = df.drop(df.index[df_index]).reset_index(drop=True)
                if "순번" in df.columns:
                    key_col_name = "학번" if sheet_name == GROUP_COUNSELING_SHEET else "이름"
                    seq_counter = 1
                    for idx, row in df.iterrows():
                        if str(row["순번"]).strip() == "예시":
                            continue
                        if str(row.get(key_col_name, "")).strip():
                            df.at[idx, "순번"] = str(seq_counter)
                            seq_counter += 1
                
                # 메모리 데이터프레임 상담회기 재정렬
                df = self._resequence_sessions_in_df(df, sheet_name, target_name, target_sid)
                self.data_frames[sheet_name] = df
                
                if os.path.exists(self.main_file_path):
                    self._last_loaded_time = os.path.getmtime(self.main_file_path)
                    self._last_checked_time = time.time()
                    
                logger.info(f"Memory Sync 완료: [{sheet_name}] row={excel_row_num} (df_index={df_index}) 삭제 및 순번 재정렬")
                return True, None

            except PermissionError:
                return False, f"'{os.path.basename(self.main_file_path)}' 파일이 다른 프로그램에서 열려있어 삭제할 수 없습니다.\n파일을 닫고 다시 시도해주세요."
            except Exception as e:
                logger.error(f"delete_excel_row 에러: {e}")
                self.restore_latest_backup()
                return False, str(e)
            finally:
                if workbook is not None:
                    try:
                        workbook.close()
                    except Exception:
                        pass
                if os.path.exists(temp_file_path):
                    try:
                        os.remove(temp_file_path)
                    except Exception:
                        pass

    def update_student_info(self, old_name, old_student_id, new_name, new_student_id, grade, gender):
        """
        학생의 개인정보(이름, 학번, 학년, 성별)를 모든 시트의 해당 레코드에서 일관되게 수정합니다.
        메모리 데이터프레임을 수정한 뒤 save_all_data_to_excel()을 호출하여 원자적으로 디스크에 반영합니다.
        """
        with self.lock:
            # 1. 엑셀 최신 데이터 확인 및 로드
            self.check_and_reload()

            normalized_old_id = old_student_id.strip().replace(".0", "")
            normalized_new_id = new_student_id.strip().replace(".0", "")
            normalized_old_name = old_name.strip()
            normalized_new_name = new_name.strip()
            normalized_gender = gender.strip()

            # 학년 포맷 정리 ("2" -> "2학년", "혼합" -> "혼합")
            grade_formatted = grade.strip()
            if grade_formatted and grade_formatted.isdigit():
                grade_formatted = f"{grade_formatted}학년"

            # 2. 이미 다른 학생이 해당 이름과 학번 조합을 사용 중인지 검증 (중복 차단)
            if (normalized_old_name != normalized_new_name) or (normalized_old_id != normalized_new_id):
                for s_name in SHEET_NAMES:
                    if s_name == GROUP_COUNSELING_SHEET:
                        continue
                    df = self.data_frames.get(s_name)
                    if df is not None and not df.empty:
                        for _, row in df.iterrows():
                            if str(row.get("순번", "")).strip() == "예시":
                                continue
                            r_name = str(row.get("이름", "")).strip()
                            r_shadow_sid = str(row.get("학번", "")).strip().replace(".0", "")
                            if r_name == normalized_new_name and r_shadow_sid == normalized_new_id:
                                # oldName/oldStudentId 에 해당하지 않는 타인 매칭 시 충돌 처리
                                if not (r_name == normalized_old_name and r_shadow_sid == normalized_old_id):
                                    return False, f"이미 학번 {normalized_new_id}(이름: {normalized_new_name})을 사용하는 학생이 존재합니다. 중복될 수 없습니다."

            modified = False

            # 3. 각 시트 순회하며 업데이트 수행
            for sheet_name in SHEET_NAMES:
                df = self.data_frames.get(sheet_name)
                if df is None or df.empty:
                    continue

                if sheet_name == GROUP_COUNSELING_SHEET:
                    # 집단상담 시트 처리 (학번 컬럼 존재)
                    if "학번" in df.columns:
                        for idx, row in df.iterrows():
                            if str(row.get("순번", "")).strip() == "예시":
                                continue
                            
                            val = str(row["학번"]).strip().replace(".0", "")
                            if not val:
                                continue

                            sids = [s.strip() for s in val.split(",") if s.strip()]
                            if normalized_old_id in sids:
                                # 학번 변경 적용
                                new_sids = [normalized_new_id if s == normalized_old_id else s for s in sids]
                                df.at[idx, "학번"] = ", ".join(new_sids)
                                modified = True

                                # 단일 학생일 경우에만 학년, 성별 업데이트
                                if len(sids) == 1:
                                    if "학년" in df.columns:
                                        df.at[idx, "학년"] = grade_formatted
                                    if "성별" in df.columns:
                                        df.at[idx, "성별"] = normalized_gender
                else:
                    # 개인상담, 보호자상담, 교원자문, 의뢰 시트 처리
                    for idx, row in df.iterrows():
                        if str(row.get("순번", "")).strip() == "예시":
                            continue
                        
                        row_name = str(row.get("이름", "")).strip()
                        row_sid = str(row.get("학번", "")).strip().replace(".0", "")

                        if row_name == normalized_old_name and row_sid == normalized_old_id:
                            df.at[idx, "이름"] = normalized_new_name
                            df.at[idx, "학번"] = normalized_new_id
                            if "학년" in df.columns:
                                df.at[idx, "학년"] = grade_formatted
                            if "성별" in df.columns:
                                df.at[idx, "성별"] = normalized_gender
                            modified = True

            if modified:
                # 변경 사항 저장 (Safe Save)
                success, err = self.save_all_data_to_excel()
                if not success:
                    return False, err
                return True, None
            else:
                return False, "수정 대상 학생 데이터를 엑셀에서 찾지 못했거나 변경사항이 없습니다."

    def delete_student_info(self, name, student_id):
        """
        학생의 모든 상담 기록(개인상담, 보호자상담, 교원자문, 의뢰)을 삭제합니다.
        집단상담 대장은 역사성 보존 및 다수 참여 특성상 삭제 대상에서 제외합니다.
        삭제 전 선제적으로 수동 백업을 1회 강제 수행합니다.
        """
        with self.lock:
            # 1. 삭제 전 안전 예비 수동 백업 수행
            backup_success, backup_result = self.save_backup_excel()
            if not backup_success:
                logger.error(f"학생 삭제 전 자동 백업 실패: {backup_result}")
                return False, f"삭제 전 선제 백업본 저장 실패로 인해 작업을 진행하지 못했습니다. 원인: {backup_result}"

            # 2. 엑셀 최신 데이터 확인 및 로드
            self.check_and_reload()

            normalized_id = student_id.strip().replace(".0", "")
            normalized_name = name.strip()
            modified = False

            # 집단상담을 제외한 타겟 시트 순회
            target_sheets = [s for s in SHEET_NAMES if s != GROUP_COUNSELING_SHEET]

            for sheet_name in target_sheets:
                df = self.data_frames.get(sheet_name)
                if df is None or df.empty:
                    continue

                indices_to_drop = []
                for idx, row in df.iterrows():
                    if str(row.get("순번", "")).strip() == "예시":
                        continue
                    
                    row_name = str(row.get("이름", "")).strip()
                    row_sid = str(row.get("학번", "")).strip().replace(".0", "")

                    if row_name == normalized_name and row_sid == normalized_id:
                        indices_to_drop.append(idx)
                        modified = True

                if indices_to_drop:
                    # 행 드롭 및 인덱스 리셋
                    df = df.drop(indices_to_drop).reset_index(drop=True)
                    
                    # 순번 재정렬
                    if "순번" in df.columns:
                        seq_counter = 1
                        key_col_name = "학번" if sheet_name == GROUP_COUNSELING_SHEET else "이름"
                        for idx, row in df.iterrows():
                            if str(row["순번"]).strip() == "예시":
                                continue
                            if str(row.get(key_col_name, "")).strip():
                                df.at[idx, "순번"] = str(seq_counter)
                                seq_counter += 1
                    
                    self.data_frames[sheet_name] = df

            if modified:
                # 엑셀 파일 저장
                success, err = self.save_all_data_to_excel()
                if not success:
                    return False, err
                return True, None
            else:
                return False, "삭제 대상 학생 데이터를 엑셀에서 찾지 못했습니다."

    def _ensure_template_initialized(self):
        """기본 템플릿 파일이 없으면 생성합니다."""
        if not os.path.exists(self.main_file_path):
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            for s_name in SHEET_NAMES:
                wb.create_sheet(title=s_name)
            wb.save(self.main_file_path)

    def _sync_headers(self, worksheet, df_to_append, border, alignment):
        """데이터프레임의 열 목록과 엑셀 시트의 헤더를 동기화하고 스타일을 입힙니다."""
        real_max = find_real_max_row(worksheet)
        if real_max == 0:
            worksheet.delete_rows(1, worksheet.max_row)
            headers_list = list(df_to_append.columns)
            worksheet.append(headers_list)
            for cell in worksheet[1]:
                cell.alignment = alignment
                cell.border = border
            real_max = 1

        headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
        for col_name in df_to_append.columns:
            if col_name not in headers:
                new_col_idx = worksheet.max_column + 1
                header_cell = worksheet.cell(row=1, column=new_col_idx, value=col_name)
                header_cell.alignment = alignment
                header_cell.border = border
                headers[col_name] = new_col_idx
        return headers, real_max

    def _get_next_sequence_number(self, worksheet, seq_col_idx, real_max):
        """엑셀 시트에서 현재 최대 순번을 연산하고 다음 순번을 생성합니다."""
        max_seq = 0
        for r in range(2, real_max + 1):
            val = worksheet.cell(row=r, column=seq_col_idx).value
            try:
                max_seq = max(max_seq, int(val))
            except (ValueError, TypeError):
                pass
        return str(max_seq + 1)

    def append_new_row_to_excel(self, sheet_name, df_to_append):
        """지정된 데이터프레임의 행을 순번에 맞춰 엑셀 파일에 추가하고, 메모리 데이터프레임도 동기화합니다."""
        with self.lock:
            if not self.main_file_path:
                return False, "저장할 파일 경로가 지정되지 않았습니다."

            # 신규 추가 시 UUID 컬럼 및 값 자동 세팅 보장
            if RECORD_ID_COL not in df_to_append.columns:
                df_to_append[RECORD_ID_COL] = ""
            for idx, row in df_to_append.iterrows():
                if not str(row[RECORD_ID_COL]).strip():
                    df_to_append.at[idx, RECORD_ID_COL] = uuid.uuid4().hex

            ext = os.path.splitext(self.main_file_path)[1]
            target_dir = os.path.dirname(self.main_file_path) or "."
            temp_file = tempfile.NamedTemporaryFile(dir=target_dir, delete=False, suffix=ext)
            temp_file_path = temp_file.name
            temp_file.close()

            workbook = None
            try:
                thin_border = Border(left=Side(style='thin'), 
                                     right=Side(style='thin'), 
                                     top=Side(style='thin'), 
                                     bottom=Side(style='thin'))
                center_alignment_with_wrap = Alignment(horizontal='center', 
                                                       vertical='center', 
                                                       wrap_text=True)

                self._ensure_template_initialized()

                shutil.copy2(self.main_file_path, temp_file_path)
                workbook = openpyxl.load_workbook(temp_file_path)
                worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(title=sheet_name)

                headers, real_max = self._sync_headers(worksheet, df_to_append, thin_border, center_alignment_with_wrap)

                target_row_idx = None
                key_column_for_check = '학번' if sheet_name == GROUP_COUNSELING_SHEET else '이름'

                if '순번' in headers and key_column_for_check in headers:
                    seq_col_idx = headers['순번']
                    key_col_idx = headers[key_column_for_check]
                    target_row_idx = find_empty_row_by_key(worksheet, seq_col_idx, key_col_idx, sheet_name)
                
                is_new_row = False
                if target_row_idx is None:
                    target_row_idx = real_max + 1
                    is_new_row = True

                row_data_dict = df_to_append.iloc[0].to_dict()
                sanitized_row = {}
                
                if is_new_row:
                    ref_row = target_row_idx - 1 if target_row_idx - 1 >= 2 else 2
                    if worksheet.row_dimensions[ref_row].height is not None:
                        worksheet.row_dimensions[target_row_idx].height = worksheet.row_dimensions[ref_row].height

                for col_name, value in row_data_dict.items():
                    if col_name in headers:
                        col_idx = headers[col_name]
                        cell = worksheet.cell(row=target_row_idx, column=col_idx)
                        s_val = self.sanitize_value(value)
                        
                        if col_name == '순번' and cell.value is not None and str(cell.value).strip() != "" and (value is None or str(value).strip() == ""):
                            s_val = str(cell.value).strip()
                            
                        if is_new_row:
                            ref_cell = worksheet.cell(row=ref_row, column=col_idx)
                            clone_cell_style(ref_cell, cell)
                            
                        cell.value = s_val
                        sanitized_row[col_name] = s_val

                if is_new_row and '순번' in headers and ('순번' not in sanitized_row or sanitized_row['순번'] == ''):
                    new_seq = self._get_next_sequence_number(worksheet, headers['순번'], real_max)
                    worksheet.cell(row=target_row_idx, column=headers['순번']).value = new_seq
                    sanitized_row['순번'] = new_seq

                # 유효성 검사 및 자동 필터 범위 재조정
                new_max = max(real_max, target_row_idx)
                update_data_validation_ranges(worksheet, real_max, new_max)
                update_auto_filter_range(worksheet, new_max)

                # _record_id 컬럼 숨김 처리
                for cell in worksheet[1]:
                    if cell.value == RECORD_ID_COL:
                        from openpyxl.utils import get_column_letter
                        worksheet.column_dimensions[get_column_letter(cell.column)].hidden = True
                        break

                # 상담회기 재정렬
                target_name = str(row_data_dict.get("이름", "")).strip()
                target_sid = str(row_data_dict.get("학번", "")).strip()
                self._resequence_sessions_in_worksheet(worksheet, sheet_name, target_name, target_sid)

                workbook.save(temp_file_path)
                workbook.close()
                workbook = None

                success, err = self.swap_temp_and_original(temp_file_path, self.main_file_path)
                if not success:
                    self.restore_latest_backup()
                    return False, err

                df = self.data_frames[sheet_name]
                df_idx = target_row_idx - 2
                
                while len(df) <= df_idx:
                    empty_row = {col: "" for col in df.columns}
                    df = pd.concat([df, pd.DataFrame([empty_row])], ignore_index=True)
                
                for col, val in sanitized_row.items():
                    df.at[df_idx, col] = val

                # 메모리 데이터프레임 상담회기 재정렬
                df = self._resequence_sessions_in_df(df, sheet_name, target_name, target_sid)
                self.data_frames[sheet_name] = df
                
                if os.path.exists(self.main_file_path):
                    self._last_loaded_time = os.path.getmtime(self.main_file_path)
                    self._last_checked_time = time.time()
                    
                logger.info(f"Memory Sync 완료: [{sheet_name}] row={target_row_idx} 추가/작성")

                return True, None

            except PermissionError:
                return False, f"'{os.path.basename(self.main_file_path)}' 파일이 다른 프로그램에서 열려있어 저장할 수 없습니다.\n파일을 닫고 다시 시도해주세요."
            except Exception as e:
                logger.error(f"append_new_row_to_excel 에러: {e}")
                self.restore_latest_backup()
                return False, str(e)
            finally:
                if workbook is not None:
                    try:
                        workbook.close()
                    except Exception:
                        pass
                if os.path.exists(temp_file_path):
                    try:
                        os.remove(temp_file_path)
                    except Exception:
                        pass

    def save_backup_excel(self):
        """메인 파일을 사용자의 '내 문서/상담일지 백업 파일' 폴더에 날짜 포함 파일명으로 복사하여 저장합니다."""
        with self.lock:
            if not self.main_file_path or not os.path.exists(self.main_file_path):
                return False, "저장된 원본 파일이 없어 백업할 수 없습니다."
    
            try:
                backup_dir = get_user_backup_path()
                ensure_directory_exists(backup_dir)
    
                backup_date = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
                file_name = f"상담일지({backup_date}).xlsx"
                destination_path = os.path.join(backup_dir, file_name)
    
                shutil.copy2(self.main_file_path, destination_path)
                logger.info(f"수동 백업 생성 완료: {destination_path}")
                return True, (file_name, backup_dir)
            except Exception as e:
                logger.error(f"save_backup_excel 에러: {e}")
                return False, str(e)

    def _resequence_sessions_in_worksheet(self, worksheet, sheet_name, target_name, target_sid):
        """지정된 학생의 상담회기 컬럼을 날짜 오름차순 기준으로 다시 1부터 순차적으로 정렬하여 저장합니다.
           실제 행의 위치는 그대로 둔 채 '상담회기' 값만 변경합니다."""
        headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
        if "상담회기" not in headers:
            return
            
        counsel_col_idx = headers["상담회기"]
        date_col_idx = headers.get("*상담일자")
        if not date_col_idx:
            return

        student_rows = []
        real_max = find_real_max_row(worksheet)
        seq_col_idx = headers.get("순번")
        
        for r in range(2, real_max + 1):
            if seq_col_idx:
                if str(worksheet.cell(row=r, column=seq_col_idx).value).strip() == "예시":
                    continue
            
            match = False
            if sheet_name == GROUP_COUNSELING_SHEET:
                if "학번" in headers and target_sid:
                    sid_val = str(worksheet.cell(row=r, column=headers["학번"]).value).strip().replace(".0", "")
                    match = (target_sid == sid_val or target_sid in [s.strip() for s in sid_val.split(",")])
            else:
                if "이름" in headers and target_name:
                    name_val = str(worksheet.cell(row=r, column=headers["이름"]).value).strip()
                    match = (name_val == target_name)
                    if match and "학번" in headers and target_sid:
                        sid_val = str(worksheet.cell(row=r, column=headers["학번"]).value).strip().replace(".0", "")
                        match = (sid_val == target_sid)
            
            if match:
                date_val = str(worksheet.cell(row=r, column=date_col_idx).value).strip()
                student_rows.append({
                    "row_num": r,
                    "date": date_val
                })
                
        student_rows.sort(key=lambda x: (x["date"], x["row_num"]))
        
        for idx, item in enumerate(student_rows):
            worksheet.cell(row=item["row_num"], column=counsel_col_idx).value = str(idx + 1)

    def _resequence_sessions_in_df(self, df, sheet_name, target_name, target_sid):
        """메모리 DataFrame의 상담회기 컬럼을 날짜 오름차순 기준으로 다시 1부터 순차적으로 정렬합니다."""
        if "상담회기" not in df.columns or "*상담일자" not in df.columns:
            return df
            
        student_indices = []
        for idx, row in df.iterrows():
            if str(row.get("순번", "")).strip() == "예시":
                continue
                
            match = False
            if sheet_name == GROUP_COUNSELING_SHEET:
                if "학번" in df.columns and target_sid:
                    sid_val = str(row["학번"]).strip().replace(".0", "")
                    match = (target_sid == sid_val or target_sid in [s.strip() for s in sid_val.split(",")])
            else:
                if "이름" in df.columns and target_name:
                    name_val = str(row["이름"]).strip()
                    match = (name_val == target_name)
                    if match and "학번" in df.columns and target_sid:
                        sid_val = str(row["학번"]).strip().replace(".0", "")
                        match = (sid_val == target_sid)
                        
            if match:
                student_indices.append({
                    "df_idx": idx,
                    "date": str(row["*상담일자"]).strip()
                })
                
        student_indices.sort(key=lambda x: (x["date"], x["df_idx"]))
        
        for idx, item in enumerate(student_indices):
            df.at[item["df_idx"], "상담회기"] = str(idx + 1)
            
        return df
