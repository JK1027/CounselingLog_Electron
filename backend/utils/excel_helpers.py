import openpyxl
from openpyxl.styles import Alignment, Border, Side

def find_real_max_row(worksheet):
    """openpyxl의 max_row 오류를 방지하기 위해 실제 값이 있는 마지막 행 번호를 찾습니다."""
    for row in range(worksheet.max_row, 0, -1):
        if any(worksheet.cell(row=row, column=col).value is not None for col in range(1, worksheet.max_column + 1)):
            return row
    return 0

def find_empty_row_by_key(worksheet, seq_col_idx, key_col_idx, sheet_name):
    """순번은 지정되었으나 이름/학번(Key)이 비어있는 빈 행의 번호를 반환합니다."""
    real_max = find_real_max_row(worksheet)
    valid_rows = []
    for r in range(2, real_max + 1):
        seq_val = worksheet.cell(row=r, column=seq_col_idx).value
        try:
            valid_rows.append({'row_num': r, 'seq': int(seq_val)})
        except (ValueError, TypeError):
            continue
    
    valid_rows.sort(key=lambda x: x['seq'])

    for row_data in valid_rows:
        r_idx = row_data['row_num']
        key_val = worksheet.cell(row=r_idx, column=key_col_idx).value
        if key_val is None or str(key_val).strip() == "":
            return r_idx
    return None

def apply_excel_formatting(worksheet):
    """주어진 워크시트에 표준 서식(센터 정렬, 테두리)을 적용하고, _record_id 컬럼을 숨김 처리합니다."""
    from openpyxl.utils import get_column_letter
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    center_alignment_with_wrap = Alignment(
        horizontal='center', 
        vertical='center', 
        wrap_text=True
    )
    
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment_with_wrap
            cell.border = thin_border

    # _record_id 컬럼 숨김 처리
    for cell in worksheet[1]:
        if cell.value == "_record_id":
            col_letter = get_column_letter(cell.column)
            worksheet.column_dimensions[col_letter].hidden = True
            break

from copy import copy
from openpyxl.utils import coordinate_to_tuple, get_column_letter

def clone_cell_style(src_cell, dst_cell):
    """원래 셀의 스타일 객체들을 대상 셀로 안전하게 복제합니다."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)

def update_data_validation_ranges(worksheet, old_max, new_max):
    """행 수가 증가했을 때, 기존 데이터 유효성 검사 범위 중 old_max로 끝나는 주소를 new_max로 확장합니다."""
    if old_max >= new_max or old_max <= 1:
        return
    
    for dv in worksheet.data_validations.dataValidation:
        sqref_parts = []
        updated = False
        for range_str in str(dv.sqref).split():
            if ":" in range_str:
                start, end = range_str.split(":")
                try:
                    end_row, end_col = coordinate_to_tuple(end)
                    if end_row == old_max:
                        col_letter = get_column_letter(end_col)
                        new_end = f"{col_letter}{new_max}"
                        sqref_parts.append(f"{start}:{new_end}")
                        updated = True
                    else:
                        sqref_parts.append(range_str)
                except Exception:
                    sqref_parts.append(range_str)
            else:
                try:
                    row_num, col_num = coordinate_to_tuple(range_str)
                    if row_num == old_max:
                        col_letter = get_column_letter(col_num)
                        sqref_parts.append(f"{range_str}:{col_letter}{new_max}")
                        updated = True
                    else:
                        sqref_parts.append(range_str)
                except Exception:
                    sqref_parts.append(range_str)
        if updated:
            dv.sqref = " ".join(sqref_parts)

def update_auto_filter_range(worksheet, new_max):
    """자동 필터 영역이 지정되어 있는 경우, 필터 끝 범위 행을 new_max로 재설정합니다."""
    if worksheet.auto_filter.ref:
        ref = worksheet.auto_filter.ref
        if ":" in ref:
            start, end = ref.split(":")
            try:
                end_row, end_col = coordinate_to_tuple(end)
                col_letter = get_column_letter(end_col)
                worksheet.auto_filter.ref = f"{start}:{col_letter}{new_max}"
            except Exception:
                pass

