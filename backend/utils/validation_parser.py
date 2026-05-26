import os
import copy
import openpyxl
from openpyxl.utils.cell import get_column_letter, range_boundaries
from backend.utils.logger import logger
from backend.core.constants import SHEET_NAMES

MAX_OPTIONS = 200

def parse_validation_options(file_path, default_options):
    """
    엑셀 파일의 각 시트별로 '*상담구분' 열에 걸려있는 데이터 유효성 검사 목록을 분석하여 반환합니다.
    분석 실패 시 default_options의 값을 활용합니다.
    """
    # 원본 오염 방지를 위해 deepcopy 적용
    options_map = {}
    for sheet_name, opts in default_options.items():
        options_map[sheet_name] = {
            "source": "fallback",
            "options": copy.deepcopy(opts)
        }
    
    # 집단상담 기본값 보정 (설정 파일에 누락된 경우 대비)
    group_sheet_name = "집단상담(또래상담, 학급별 집단)"
    if group_sheet_name not in options_map:
        options_map[group_sheet_name] = {
            "source": "fallback",
            "options": ['정신건강', '진로', '학업', '대인관계', '기타']
        }
        
    if not file_path or not os.path.exists(file_path):
        return options_map
        
    workbook = None
    try:
        # data_only=False로 로드하여 Formula 원본 추출 보장
        workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=False)
        
        for sheet_name in SHEET_NAMES:
            if sheet_name not in workbook.sheetnames:
                continue
            
            worksheet = workbook[sheet_name]
            headers = [cell.value for cell in worksheet[1]]
            
            target_col_name = "*상담구분"
            if target_col_name not in headers:
                continue
                
            col_idx = headers.index(target_col_name) + 1
            
            found_options = []
            for dv in worksheet.data_validations.dataValidation:
                match_column = False
                for range_str in str(dv.sqref).split():
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(range_str)
                        if min_col <= col_idx <= max_col:
                            match_column = True
                            break
                    except Exception as e:
                        logger.debug(f"[{sheet_name}] sqref boundaries parsing error for {range_str}: {e}")
                
                if match_column and dv.formula1:
                    formula = dv.formula1
                    
                    # Case 1: 직접 입력 목록 형태 (예: '"값1,값2,값3"')
                    if formula.startswith('"') and formula.endswith('"'):
                        found_options = [opt.strip() for opt in formula[1:-1].split(",") if opt.strip()]
                        break
                    
                    # Case 2: 셀 범위 참조 형태 (예: 'Sheet2!$A$1:$A$10' 혹은 '=Sheet2!$A$1:$A$10')
                    ref = formula[1:] if formula.startswith('=') else formula
                    ref_sheet = sheet_name
                    ref_range = ref
                    
                    if '!' in ref:
                        ref_sheet, ref_range = ref.split('!', 1)
                        if ref_sheet.startswith("'") and ref_sheet.endswith("'"):
                            ref_sheet = ref_sheet[1:-1]
                    
                    ref_range = ref_range.replace('$', '')
                    
                    # 전체 열 참조(A:A) 전처리: NoneType 방지 및 폭주 제어
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(ref_range)
                        if min_row is None:
                            min_row = 2
                        if max_row is None:
                            max_row = min_row + MAX_OPTIONS
                        
                        # 범위 크기가 MAX_OPTIONS를 넘어갈 경우 재조정
                        if max_row - min_row > MAX_OPTIONS:
                            max_row = min_row + MAX_OPTIONS
                            
                        ref_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                    except Exception as e:
                        logger.debug(f"[{sheet_name}] range pre-process error for {ref_range}: {e}")
                    
                    if ref_sheet in workbook.sheetnames:
                        try:
                            ref_ws = workbook[ref_sheet]
                            cells = ref_ws[ref_range]
                            opts = []
                            if isinstance(cells, tuple):
                                for row in cells:
                                    for cell in row:
                                        if cell.value is not None:
                                            opts.append(str(cell.value).strip())
                            else:
                                if cells.value is not None:
                                    opts.append(str(cells.value).strip())
                            if opts:
                                found_options = opts
                                break
                        except Exception as e:
                            logger.debug(f"[{sheet_name}] Reading range cells {ref_range} failed: {e}")
                                    
                    # Case 3: 이름 정의 범위 형태 (Named Range)
                    if ref in workbook.defined_names:
                        try:
                            destinations = workbook.defined_names[ref].destinations
                            opts = []
                            for dest_sheet, dest_range in destinations:
                                if dest_sheet in workbook.sheetnames:
                                    dest_range = dest_range.replace('$', '')
                                    
                                    # 전체 열 참조 전처리
                                    min_col, min_row, max_col, max_row = range_boundaries(dest_range)
                                    if min_row is None:
                                        min_row = 2
                                    if max_row is None:
                                        max_row = min_row + MAX_OPTIONS
                                        
                                    if max_row - min_row > MAX_OPTIONS:
                                        max_row = min_row + MAX_OPTIONS
                                    dest_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                                        
                                    ref_ws = workbook[dest_sheet]
                                    cells = ref_ws[dest_range]
                                    if isinstance(cells, tuple):
                                        for row in cells:
                                            for cell in row:
                                                if cell.value is not None:
                                                    opts.append(str(cell.value).strip())
                                    else:
                                        if cells.value is not None:
                                            opts.append(str(cells.value).strip())
                            if opts:
                                found_options = opts
                                break
                        except Exception as e:
                            logger.debug(f"[{sheet_name}] Resolving Named Range {ref} failed: {e}")
                            
            if found_options:
                unique_opts = list(dict.fromkeys([opt for opt in found_options if opt.strip()]))
                if len(unique_opts) > MAX_OPTIONS:
                    unique_opts = unique_opts[:MAX_OPTIONS]
                options_map[sheet_name] = {
                    "source": "excel",
                    "options": unique_opts
                }
                logger.info(f"엑셀 시트 [{sheet_name}]로부터 동적 드롭다운 목록 추출 성공 (개수: {len(unique_opts)}): {unique_opts}")
                
        return options_map
    except Exception as e:
        logger.error(f"엑셀 데이터 유효성 검사 파싱 실패 (config fallback 작동): {e}")
        return options_map
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
