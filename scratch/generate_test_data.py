"""
상담일지 프로그램 기능 테스트를 위한 엑셀 테스트 파일 생성 스크립트
- 50명의 학생 데이터
- 5개 시트 전체에 걸친 다양한 상담 기록
- 초성 검색, 학번 검색, 정렬, 통계, 회기 관리, 필터링 등 테스트 가능
"""

import os
import uuid
import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────
# 1. 시트 스키마 정의 (constants.py 기준)
# ─────────────────────────────────────────────────────
RECORD_ID_COL = "_record_id"

SHEET_SCHEMAS = {
    "개인상담": [
        "순번", "학번", "이름", "상담시간", "기타 및 특이사항",
        "*상담분류", "*Wee클래스", "*대분류", "*중분류", "*상담구분", "*상담인원", "*학년도", "*상담일자",
        "학년", "성별", "*상담제목", "*상담내용", "상담내용(상세)",
        "*상담시간(시)", "*상담시간(분)", "*상담사소속", "*상담매체구분", "상담회기", RECORD_ID_COL
    ],
    "집단상담(또래상담, 학급별 집단)": [
        "순번", "학번", "프로그램명", "상담시간", "상담회기", "목표", "기타 및 특이사항",
        "*상담분류", "*Wee클래스", "*대분류", "*중분류", "*상담구분", "*상담인원", "*학년도", "*상담일자",
        "학년", "성별", "*상담제목", "*상담내용", "상담내용(상세)",
        "*상담시간(시)", "*상담시간(분)", "*상담사소속", "*상담매체구분", RECORD_ID_COL
    ],
    "보호자상담": [
        "순번", "학번", "이름", "상담시간", "기타 및 특이사항",
        "*상담분류", "*Wee클래스", "*대분류", "*중분류", "*상담구분", "*상담인원", "*학년도", "*상담일자",
        "학년", "성별", "*상담제목", "*상담내용", "상담내용(상세)",
        "*상담시간(시)", "*상담시간(분)", "*상담사소속", "*상담매체구분", RECORD_ID_COL
    ],
    "교원자문": [
        "순번", "학번", "이름", "상담시간", "기타 및 특이사항",
        "*상담분류", "*Wee클래스", "*대분류", "*중분류", "*상담구분", "*상담인원", "*학년도", "*상담일자",
        "학년", "성별", "*상담제목", "*상담내용", "상담내용(상세)",
        "*상담시간(시)", "*상담시간(분)", "*상담사소속", "*상담매체구분", RECORD_ID_COL
    ],
    "의뢰(정서행동의뢰, 자문의 의뢰 등)": [
        "순번", "학번", "이름", "상담시간", "기타 및 특이사항",
        "*상담분류", "*Wee클래스", "*대분류", "*중분류", "*상담구분", "*상담인원", "*학년도", "*상담일자",
        "학년", "성별", "*상담제목", "*상담내용", "상담내용(상세)",
        "*상담시간(시)", "*상담시간(분)", "*상담사소속", "*상담매체구분", RECORD_ID_COL
    ]
}

# ─────────────────────────────────────────────────────
# 2. 50명 학생 데이터 정의
# ─────────────────────────────────────────────────────
# 다양한 테스트 시나리오를 커버하기 위해 설계:
# - 초성 검색 테스트: 같은 성(김, 이, 박 등) 다수 포함
# - 학번 4자리/5자리: 초등학교(4자리) 기준
# - 학년: 1~6학년 골고루 분포
# - 반: 1~3반
# - 성별: 남/여 균등 분포
# - 동명이인 테스트: 같은 이름 다른 학번

STUDENTS = [
    # === 1학년 (grade=1) ===
    {"name": "김민수", "sid": "1101", "grade": "1", "gender": "남", "ban": "1"},   # 1
    {"name": "김서연", "sid": "1102", "grade": "1", "gender": "여", "ban": "1"},   # 2
    {"name": "이준호", "sid": "1103", "grade": "1", "gender": "남", "ban": "1"},   # 3
    {"name": "박지은", "sid": "1201", "grade": "1", "gender": "여", "ban": "2"},   # 4
    {"name": "최유진", "sid": "1202", "grade": "1", "gender": "여", "ban": "2"},   # 5
    {"name": "정하윤", "sid": "1203", "grade": "1", "gender": "여", "ban": "2"},   # 6
    {"name": "강도윤", "sid": "1301", "grade": "1", "gender": "남", "ban": "3"},   # 7
    {"name": "윤서현", "sid": "1302", "grade": "1", "gender": "여", "ban": "3"},   # 8
    
    # === 2학년 (grade=2) ===
    {"name": "김민수", "sid": "2101", "grade": "2", "gender": "남", "ban": "1"},   # 9 동명이인!
    {"name": "이서아", "sid": "2102", "grade": "2", "gender": "여", "ban": "1"},   # 10
    {"name": "한지우", "sid": "2103", "grade": "2", "gender": "여", "ban": "1"},   # 11
    {"name": "조현우", "sid": "2201", "grade": "2", "gender": "남", "ban": "2"},   # 12
    {"name": "서예린", "sid": "2202", "grade": "2", "gender": "여", "ban": "2"},   # 13
    {"name": "임도현", "sid": "2203", "grade": "2", "gender": "남", "ban": "2"},   # 14
    {"name": "오수빈", "sid": "2301", "grade": "2", "gender": "여", "ban": "3"},   # 15
    {"name": "황재민", "sid": "2302", "grade": "2", "gender": "남", "ban": "3"},   # 16

    # === 3학년 (grade=3) ===
    {"name": "김하은", "sid": "3101", "grade": "3", "gender": "여", "ban": "1"},   # 17
    {"name": "이도윤", "sid": "3102", "grade": "3", "gender": "남", "ban": "1"},   # 18
    {"name": "박수아", "sid": "3103", "grade": "3", "gender": "여", "ban": "1"},   # 19
    {"name": "최건우", "sid": "3201", "grade": "3", "gender": "남", "ban": "2"},   # 20
    {"name": "정민서", "sid": "3202", "grade": "3", "gender": "여", "ban": "2"},   # 21
    {"name": "송지호", "sid": "3203", "grade": "3", "gender": "남", "ban": "2"},   # 22
    {"name": "나윤아", "sid": "3301", "grade": "3", "gender": "여", "ban": "3"},   # 23
    {"name": "문태영", "sid": "3302", "grade": "3", "gender": "남", "ban": "3"},   # 24

    # === 4학년 (grade=4) ===
    {"name": "김지안", "sid": "4101", "grade": "4", "gender": "여", "ban": "1"},   # 25
    {"name": "이건호", "sid": "4102", "grade": "4", "gender": "남", "ban": "1"},   # 26
    {"name": "박서준", "sid": "4103", "grade": "4", "gender": "남", "ban": "1"},   # 27
    {"name": "장하린", "sid": "4201", "grade": "4", "gender": "여", "ban": "2"},   # 28
    {"name": "신우진", "sid": "4202", "grade": "4", "gender": "남", "ban": "2"},   # 29
    {"name": "고은서", "sid": "4203", "grade": "4", "gender": "여", "ban": "2"},   # 30
    {"name": "유시현", "sid": "4301", "grade": "4", "gender": "남", "ban": "3"},   # 31
    {"name": "배다인", "sid": "4302", "grade": "4", "gender": "여", "ban": "3"},   # 32

    # === 5학년 (grade=5) ===
    {"name": "김태윤", "sid": "5101", "grade": "5", "gender": "남", "ban": "1"},   # 33
    {"name": "이하음", "sid": "5102", "grade": "5", "gender": "여", "ban": "1"},   # 34
    {"name": "박채원", "sid": "5103", "grade": "5", "gender": "여", "ban": "1"},   # 35
    {"name": "정지호", "sid": "5201", "grade": "5", "gender": "남", "ban": "2"},   # 36
    {"name": "한소율", "sid": "5202", "grade": "5", "gender": "여", "ban": "2"},   # 37
    {"name": "류민재", "sid": "5203", "grade": "5", "gender": "남", "ban": "2"},   # 38
    {"name": "양서윤", "sid": "5301", "grade": "5", "gender": "여", "ban": "3"},   # 39
    {"name": "권도현", "sid": "5302", "grade": "5", "gender": "남", "ban": "3"},   # 40

    # === 6학년 (grade=6) ===
    {"name": "김소윤", "sid": "6101", "grade": "6", "gender": "여", "ban": "1"},   # 41
    {"name": "이현준", "sid": "6102", "grade": "6", "gender": "남", "ban": "1"},   # 42
    {"name": "박나윤", "sid": "6103", "grade": "6", "gender": "여", "ban": "1"},   # 43
    {"name": "최준서", "sid": "6201", "grade": "6", "gender": "남", "ban": "2"},   # 44
    {"name": "장서아", "sid": "6202", "grade": "6", "gender": "여", "ban": "2"},   # 45
    {"name": "홍길동", "sid": "6203", "grade": "6", "gender": "남", "ban": "2"},   # 46  특이 이름
    {"name": "구미래", "sid": "6301", "grade": "6", "gender": "여", "ban": "3"},   # 47
    {"name": "탁현우", "sid": "6302", "grade": "6", "gender": "남", "ban": "3"},   # 48
    {"name": "피서진", "sid": "6303", "grade": "6", "gender": "여", "ban": "3"},   # 49  드문 성씨
    {"name": "하준영", "sid": "6304", "grade": "6", "gender": "남", "ban": "3"},   # 50
]

# ─────────────────────────────────────────────────────
# 3. 상담 유형별 옵션 (config.json 기준)
# ─────────────────────────────────────────────────────
COUNSELING_TYPES = {
    "개인상담": [
        "학업", "진로", "성격", "성", "대인관계", "가정 및 가족관계",
        "일탈 및 비행", "학교폭력 가해", "학교폭력 피해", "자해 및 자살",
        "정신건강", "컴퓨터 및 스마트폰 과사용", "정보제공", "기타"
    ],
    "보호자상담": ["학생관련상담", "교사관련상담", "학습", "기타"],
    "교원자문": ["학교학습", "사회성발달", "정서발달", "진로발달", "행동발달", "기타"],
    "의뢰": ["외부전문가에게 상담의뢰", "교내 교사에게 상담의뢰", "기타"],
}

# ─────────────────────────────────────────────────────
# 4. 상담 내용 템플릿
# ─────────────────────────────────────────────────────
COUNSELING_SUMMARIES = {
    "학업": [
        "학습 동기 저하 상담", "시험 불안 해소", "학습 계획 수립 지도",
        "과목별 학습 전략 상담", "자기주도학습 역량 강화"
    ],
    "진로": [
        "진로 탐색 상담", "진로 적성 검사 결과 해석", "꿈 발표 준비 상담",
        "직업 체험 사전 상담", "진로 희망 사항 탐색"
    ],
    "성격": [
        "자아 존중감 향상 상담", "성격 유형 이해 상담", "자기 이해 활동",
        "감정 조절 훈련", "강점 발견 프로그램"
    ],
    "대인관계": [
        "친구 관계 갈등 상담", "또래 관계 개선", "사회성 기술 훈련",
        "갈등 해결 방법 모색", "의사소통 기술 향상"
    ],
    "가정 및 가족관계": [
        "가정 내 갈등 상담", "부모-자녀 관계 개선", "가족 구조 변화 적응 지원",
        "양육 환경 관련 상담", "형제 갈등 중재"
    ],
    "학교폭력 피해": [
        "학교폭력 피해 신고 접수", "피해 학생 심리 지원", "또래 관계 회복 지원",
        "안전 계획 수립", "보호자 연계 상담"
    ],
    "학교폭력 가해": [
        "가해 행동 인식 상담", "공감 능력 향상 훈련", "분노 조절 프로그램",
        "행동 수정 계약서 작성", "재발 방지 교육"
    ],
    "자해 및 자살": [
        "자해 행동 위기 상담", "안전 계획 수립", "심리 안정 지원",
        "전문 기관 연계 상담", "추적 관찰 상담"
    ],
    "정신건강": [
        "불안 증상 상담", "우울감 호소 상담", "ADHD 관련 행동 관찰",
        "스트레스 관리 상담", "정서 안정화 활동"
    ],
    "컴퓨터 및 스마트폰 과사용": [
        "스마트폰 사용 습관 점검", "디지털 디톡스 계획 수립", "게임 과몰입 상담",
        "인터넷 윤리 교육", "미디어 리터러시 상담"
    ],
    "정보제공": [
        "상담 안내 및 정보제공", "프로그램 참여 안내", "외부 기관 정보 안내",
        "심리검사 안내", "학교 적응 정보 제공"
    ],
    "기타": [
        "전학 적응 상담", "기타 생활 상담", "일반 상담",
        "학교생활 적응 상담", "기본 면담"
    ],
    "성": [
        "성교육 관련 상담", "성인식 교육 연계", "건강한 신체 발달 상담"
    ],
    "일탈 및 비행": [
        "무단결석 관련 상담", "교내 규칙 위반 상담", "또래 비행 예방 교육",
        "생활지도 연계 상담"
    ],
    # 보호자상담
    "학생관련상담": [
        "자녀 학교생활 적응 상담", "자녀 행동 변화 상담", "학습 부진 관련 학부모 면담",
        "자녀 교우 관계 상담", "자녀 진로 상담"
    ],
    "교사관련상담": [
        "담임교사-학부모 소통 지원", "교사와의 관계 개선 상담"
    ],
    "학습": [
        "가정 학습 환경 조성 상담", "학습 지원 방법 안내", "학원 선택 관련 상담"
    ],
    # 교원자문
    "학교학습": [
        "학습 부진아 지도 방법 자문", "수업 중 집중력 향상 방안"
    ],
    "사회성발달": [
        "또래 관계 지도 방법 자문", "학급 내 갈등 해결 방안"
    ],
    "정서발달": [
        "학생 정서 불안 대처 방법", "감정 코칭 기법 자문"
    ],
    "진로발달": [
        "진로 교육 연계 방안 자문", "진로 체험 프로그램 설계"
    ],
    "행동발달": [
        "행동 문제 학생 지도 자문", "ADHD 의심 학생 관찰 지도"
    ],
    # 의뢰
    "외부전문가에게 상담의뢰": [
        "외부 전문 상담 기관 의뢰", "정신건강의학과 연계 의뢰", "아동보호전문기관 의뢰"
    ],
    "교내 교사에게 상담의뢰": [
        "담임교사 연계 의뢰", "보건교사 연계 의뢰"
    ],
}

COUNSELING_DETAILS = {
    "학업": [
        "학생이 최근 성적 하락으로 인해 학습 의욕이 저하된 상태임. 함께 학습 목표를 재설정하고 단계별 학습 계획을 세워 보았음. 학생은 수학 과목에 대한 두려움이 큰 상태로, 기초부터 다시 시작하는 것에 동의함.",
        "시험 기간이 다가오면서 불안감이 높아진 상태. 호흡법과 마인드맵 기법을 활용한 학습 전략을 안내함. 시험 후 결과와 관계없이 자기 격려를 할 수 있도록 격려함.",
        "자기주도학습 습관이 형성되지 않아 매일 학습 시간을 기록하는 일지를 작성하기로 약속함. 1주일 후 점검 예정.",
    ],
    "진로": [
        "학생이 장래 희망이 아직 구체적이지 않아 다양한 직업을 탐색해보기로 함. 관심 분야 리스트를 작성하고, 관련 직업인을 인터뷰하는 과제를 제시함.",
        "진로 적성 검사 결과를 함께 해석하며, 학생의 강점 영역(언어·사회)을 확인함. 관련 진로 체험 프로그램 참여를 권유함.",
    ],
    "대인관계": [
        "같은 반 친구와의 갈등으로 학교에 오기 싫다고 호소함. 갈등 상황을 구체적으로 들어보고, 서로의 입장을 이해해보는 역할극을 진행함. 다음 상담에서 후속 확인 예정.",
        "쉬는 시간에 혼자 있는 시간이 많다며 외로움을 표현함. 관심사가 비슷한 또래와의 활동을 제안하고, 동아리 가입을 함께 알아봄.",
        "새로 전학 온 학생으로 아직 반 분위기에 적응하지 못하고 있음. 담임교사와 협력하여 짝꿍 배치를 조정하고, 또래 멘토를 배정하기로 함.",
    ],
    "가정 및 가족관계": [
        "부모님의 이혼 후 정서적으로 불안정한 모습을 보임. 감정 일기를 작성하도록 안내하고, 안전한 공간에서 감정을 표현할 수 있도록 지지함.",
        "가정에서 부모님과의 대화가 부족하다고 호소함. 가정 내 대화 시간을 만들 수 있도록 부모님께 편지를 써보기로 함.",
    ],
    "학교폭력 피해": [
        "등교 시간에 특정 학생들로부터 지속적인 언어적 괴롭힘을 받고 있다고 신고함. 즉각적인 안전 조치(교실 분리, 등하교 동행)를 취하고, 학교폭력 대책자치위원회 심의 절차를 안내함. 보호자에게 연락 완료.",
        "SNS를 통한 사이버불링 피해를 호소함. 증거 자료를 확보하고 전담 경찰관에게 연계 예정. 심리적 안정을 위한 지속 상담 필요.",
    ],
    "학교폭력 가해": [
        "또래 학생을 지속적으로 놀린 행위에 대해 상담함. 자신의 행동이 상대에게 미치는 영향을 인식할 수 있도록 공감 훈련을 진행함. 사과 편지 쓰기 과제를 부여함.",
    ],
    "자해 및 자살": [
        "팔뚝에 자해 흔적이 발견되어 긴급 상담 진행. 현재 자살 사고(思考)는 부정하나 자해 빈도가 증가 추세임. 안전 계획서를 작성하고, 보호자 동의 하에 외부 전문 기관(정신건강복지센터) 연계 예정. 주 2회 추적 상담 계획.",
        "친한 친구의 전학으로 극심한 상실감을 호소하며 '아무것도 하기 싫다'고 표현함. 자살 사고 선별 질문에서 수동적 사고('없어지고 싶다')를 확인. 1:1 안전 관찰을 시작하고 보호자에게 즉시 통보함.",
    ],
    "정신건강": [
        "수업 중 갑자기 울거나 멍하게 있는 모습이 반복적으로 관찰됨. 기본 정서 상태를 확인한 결과 최근 2주간 수면 장애, 식욕 저하 등을 호소함. 외부 전문기관 연계를 고려 중.",
        "교실에서 갑작스러운 불안 발작(심장 두근거림, 호흡 곤란)을 경험함. 응급 이완 기법을 지도하고, 보건교사와 협력하여 발작 시 대응 계획을 수립함.",
    ],
    "컴퓨터 및 스마트폰 과사용": [
        "매일 5시간 이상 스마트폰 게임을 한다고 보고함. 하루 사용 시간 기록지를 작성하고, 단계적으로 사용 시간을 줄이는 목표를 설정함. 보호자와의 협력 상담 예정.",
    ],
    "정보제공": [
        "Wee클래스 이용 안내 및 상담 절차에 대해 설명함. 학생이 상담에 대한 긍정적 인식을 갖도록 유도함.",
        "외부 청소년 상담 기관(청소년상담복지센터, 1388) 정보를 제공함.",
    ],
    "기타": [
        "전학 후 새 학교에 적응하는 과정에서 겪는 어려움에 대해 이야기를 나눔. 전반적으로 안정적인 모습이나 수학 수업에서 진도 차이로 스트레스를 받고 있음.",
        "특별한 호소 사항 없이 정기적 안부 확인 차 면담함. 전반적으로 밝고 안정적인 모습.",
    ],
    "성": [
        "성교육 수업 후 궁금한 점을 질문하러 옴. 연령에 맞는 올바른 성 지식을 안내하고, 건강한 신체 발달에 대해 설명함."
    ],
    "일탈 및 비행": [
        "무단결석 3회 발생. 학생과 면담하여 결석 사유를 확인함. 가정 내 문제(부모 부재)가 원인으로 파악됨. 지역사회 복지 기관 연계 검토 중.",
    ],
    "학생관련상담": [
        "어머니가 자녀의 학교생활 적응에 대해 걱정하여 내방함. 학교에서의 행동 관찰 내용을 공유하고, 가정에서의 양육 전략을 함께 논의함.",
        "아버지가 자녀의 교우 관계에 대해 상담을 요청함. 학교에서 친구 사귀기 프로그램에 참여시킬 것을 제안함.",
    ],
    "교사관련상담": [
        "학부모가 담임교사와의 의사소통에 어려움을 호소함. 양자 간 면담을 주선하기로 함.",
    ],
    "학습": [
        "가정에서 자녀의 학습을 어떻게 지원해야 할지 모르겠다는 학부모 상담. 연령대에 맞는 학습 지도 방법과 추천 교재를 안내함.",
    ],
    "학교학습": [
        "학습 부진 학생에 대한 효과적인 지도 방법을 담임교사에게 자문함. 개별화 학습 전략과 보상 체계 도입을 제안함.",
    ],
    "사회성발달": [
        "반에서 따돌림이 의심되는 상황에 대해 담임교사에게 관찰 및 개입 방법을 자문함.",
    ],
    "정서발달": [
        "수업 중 감정 폭발이 잦은 학생에 대해 담임교사에게 감정 코칭 기법을 안내함. 교실 내 안정 공간 마련을 제안함.",
    ],
    "진로발달": [
        "진로 수업 커리큘럼에 체험활동을 포함시키는 방안에 대해 진로 담당 교사와 논의함.",
    ],
    "행동발달": [
        "수업 중 이석, 주의산만 등 행동 문제를 보이는 학생에 대해 담임에게 행동 수정 기법을 안내함. ADHD 선별 검사를 권고함.",
    ],
    "외부전문가에게 상담의뢰": [
        "정서행동특성검사 결과 관심군으로 분류된 학생을 외부 전문 상담 기관에 의뢰함. 보호자 동의서 확보 완료. 정신건강복지센터에 연계 예정.",
    ],
    "교내 교사에게 상담의뢰": [
        "보건 관련 이슈(두통, 복통 반복 호소)로 보건교사에게 학생을 의뢰함. 심리적 요인 가능성을 함께 전달.",
    ],
}

COUNSELING_TIMES = [
    "08:30~09:10", "09:10~09:50", "10:00~10:40", "10:50~11:30",
    "11:30~12:10", "12:30~13:10", "13:10~13:50", "14:00~14:40",
    "14:50~15:30"
]

GROUP_PROGRAM_NAMES = [
    "또래관계 향상 프로그램", "자아 존중감 향상 집단상담",
    "학교폭력 예방 교육", "감정 코칭 집단상담",
    "진로 탐색 프로그램", "사회성 기술 훈련",
    "스트레스 관리 집단상담", "의사소통 기술 향상 프로그램"
]

GROUP_COUNSELING_TYPES = [
    "학업", "진로", "성격", "대인관계", "정신건강", "기타"
]

# ─────────────────────────────────────────────────────
# 5. 날짜 생성 유틸리티
# ─────────────────────────────────────────────────────
def generate_dates_in_range(start_str, end_str, count):
    """지정된 범위 내에서 무작위 날짜를 생성 (YYYYMMDD 포맷)"""
    start = datetime.strptime(start_str, "%Y%m%d")
    end = datetime.strptime(end_str, "%Y%m%d")
    delta = (end - start).days
    dates = sorted([start + timedelta(days=random.randint(0, delta)) for _ in range(count)])
    return [d.strftime("%Y%m%d") for d in dates]

# 오늘 날짜 (통계 테스트용)
TODAY = datetime.now().strftime("%Y%m%d")
YEAR = datetime.now().strftime("%Y")

# ─────────────────────────────────────────────────────
# 6. 예시 행 데이터 (첫 번째 행에 위치하는 안내 행)
# ─────────────────────────────────────────────────────
def create_example_row(schema, sheet_name):
    """예시 행 생성 - 순번이 '예시'인 안내 데이터"""
    row = {col: "" for col in schema}
    row["순번"] = "예시"
    if "학번" in row: row["학번"] = "1101"
    if "이름" in row: row["이름"] = "홍길동"
    if "상담시간" in row: row["상담시간"] = "12:30~13:10"
    row["*상담분류"] = "전문상담"
    row["*Wee클래스"] = "Wee클래스"
    if sheet_name == "개인상담":
        row["*대분류"] = "상담"
        row["*중분류"] = "개인상담"
    elif sheet_name == "집단상담(또래상담, 학급별 집단)":
        row["*대분류"] = "상담"
        row["*중분류"] = "집단상담"
        row["프로그램명"] = "또래상담"
        row["상담회기"] = "1"
        row["목표"] = "또래관계 향상"
    elif sheet_name == "보호자상담":
        row["*대분류"] = "상담"
        row["*중분류"] = "학부모상담"
    elif sheet_name == "교원자문":
        row["*대분류"] = "자문"
        row["*중분류"] = "교원자문"
    elif "의뢰" in sheet_name:
        row["*대분류"] = "의뢰"
        row["*중분류"] = "교내외의뢰"
    row["*상담구분"] = "기타"
    row["*상담인원"] = "1"
    row["*학년도"] = YEAR
    row["*상담일자"] = TODAY
    row["학년"] = "1학년"
    row["성별"] = "남"
    row["*상담제목"] = "예시 상담 제목"
    row["*상담내용"] = "예시 상담 제목"
    row["상담내용(상세)"] = "이 행은 예시 데이터입니다. 이 행을 삭제하지 마세요."
    row["*상담시간(시)"] = "0"
    row["*상담시간(분)"] = "40"
    row["*상담사소속"] = "전문상담사"
    row["*상담매체구분"] = "면담"
    if "상담회기" in row and sheet_name != "집단상담(또래상담, 학급별 집단)":
        row["상담회기"] = "1"
    row[RECORD_ID_COL] = uuid.uuid4().hex
    return row


# ─────────────────────────────────────────────────────
# 7. 개인상담 데이터 생성
# ─────────────────────────────────────────────────────
def generate_individual_counseling():
    """
    개인상담 시트 데이터 생성
    - 50명 학생 중 35명이 개인상담 기록 보유
    - 일부 학생은 다회기(2~8회), 일부는 단회기
    - 오늘 날짜 포함 (오늘 통계 테스트용)
    - 다양한 상담구분 커버
    """
    rows = []
    seq = 1
    
    # 다회기 학생들 (8~15명, 2~8회기)
    multi_session_students = random.sample(STUDENTS[:40], 12)
    # 단회기 학생들 (나머지에서 선택)
    single_session_pool = [s for s in STUDENTS if s not in multi_session_students]
    single_session_students = random.sample(single_session_pool, 15)
    
    # 오늘 상담이 있는 학생 (통계 테스트용) - 3~5명
    today_students = random.sample(STUDENTS[:30], 4)
    
    all_types_used = set()
    
    for student in multi_session_students:
        num_sessions = random.randint(2, 8)
        # 한 학생이 같은 상담구분 또는 다른 구분으로 여러 회기
        primary_type = random.choice(COUNSELING_TYPES["개인상담"])
        all_types_used.add(primary_type)
        
        dates = generate_dates_in_range(f"{YEAR}0301", f"{YEAR}0520", num_sessions)
        
        for i, date in enumerate(dates):
            # 일부 세션은 다른 상담구분
            if i > 0 and random.random() < 0.2:
                counseling_type = random.choice(COUNSELING_TYPES["개인상담"])
            else:
                counseling_type = primary_type
            
            summaries = COUNSELING_SUMMARIES.get(counseling_type, COUNSELING_SUMMARIES["기타"])
            details = COUNSELING_DETAILS.get(counseling_type, COUNSELING_DETAILS["기타"])
            
            row = {col: "" for col in SHEET_SCHEMAS["개인상담"]}
            row["순번"] = str(seq)
            row["학번"] = student["sid"]
            row["이름"] = student["name"]
            row["상담시간"] = random.choice(COUNSELING_TIMES)
            row["기타 및 특이사항"] = "" if random.random() < 0.7 else "담임교사 연계 필요"
            row["*상담분류"] = "전문상담"
            row["*Wee클래스"] = "Wee클래스"
            row["*대분류"] = "상담"
            row["*중분류"] = "개인상담"
            row["*상담구분"] = counseling_type
            row["*상담인원"] = "1"
            row["*학년도"] = date[:4]
            row["*상담일자"] = date
            row["학년"] = f"{student['grade']}학년"
            row["성별"] = student["gender"]
            row["*상담제목"] = random.choice(summaries)
            row["*상담내용"] = row["*상담제목"]
            row["상담내용(상세)"] = random.choice(details)
            row["*상담시간(시)"] = "0"
            row["*상담시간(분)"] = random.choice(["30", "40", "50"])
            row["*상담사소속"] = "전문상담사"
            row["*상담매체구분"] = random.choice(["면담", "전화", "사이버"])
            row["상담회기"] = str(i + 1)
            row[RECORD_ID_COL] = uuid.uuid4().hex
            rows.append(row)
            seq += 1
    
    for student in single_session_students:
        counseling_type = random.choice(COUNSELING_TYPES["개인상담"])
        all_types_used.add(counseling_type)
        summaries = COUNSELING_SUMMARIES.get(counseling_type, COUNSELING_SUMMARIES["기타"])
        details = COUNSELING_DETAILS.get(counseling_type, COUNSELING_DETAILS["기타"])
        
        date = generate_dates_in_range(f"{YEAR}0301", f"{YEAR}0520", 1)[0]
        
        row = {col: "" for col in SHEET_SCHEMAS["개인상담"]}
        row["순번"] = str(seq)
        row["학번"] = student["sid"]
        row["이름"] = student["name"]
        row["상담시간"] = random.choice(COUNSELING_TIMES)
        row["*상담분류"] = "전문상담"
        row["*Wee클래스"] = "Wee클래스"
        row["*대분류"] = "상담"
        row["*중분류"] = "개인상담"
        row["*상담구분"] = counseling_type
        row["*상담인원"] = "1"
        row["*학년도"] = date[:4]
        row["*상담일자"] = date
        row["학년"] = f"{student['grade']}학년"
        row["성별"] = student["gender"]
        row["*상담제목"] = random.choice(summaries)
        row["*상담내용"] = row["*상담제목"]
        row["상담내용(상세)"] = random.choice(details)
        row["*상담시간(시)"] = "0"
        row["*상담시간(분)"] = "40"
        row["*상담사소속"] = "전문상담사"
        row["*상담매체구분"] = "면담"
        row["상담회기"] = "1"
        row[RECORD_ID_COL] = uuid.uuid4().hex
        rows.append(row)
        seq += 1
    
    # 오늘 날짜 상담 추가 (오늘 통계 테스트)
    for student in today_students:
        counseling_type = random.choice(COUNSELING_TYPES["개인상담"])
        summaries = COUNSELING_SUMMARIES.get(counseling_type, COUNSELING_SUMMARIES["기타"])
        details = COUNSELING_DETAILS.get(counseling_type, COUNSELING_DETAILS["기타"])
        
        row = {col: "" for col in SHEET_SCHEMAS["개인상담"]}
        row["순번"] = str(seq)
        row["학번"] = student["sid"]
        row["이름"] = student["name"]
        row["상담시간"] = random.choice(COUNSELING_TIMES)
        row["*상담분류"] = "전문상담"
        row["*Wee클래스"] = "Wee클래스"
        row["*대분류"] = "상담"
        row["*중분류"] = "개인상담"
        row["*상담구분"] = counseling_type
        row["*상담인원"] = "1"
        row["*학년도"] = YEAR
        row["*상담일자"] = TODAY
        row["학년"] = f"{student['grade']}학년"
        row["성별"] = student["gender"]
        row["*상담제목"] = random.choice(summaries)
        row["*상담내용"] = row["*상담제목"]
        row["상담내용(상세)"] = random.choice(details)
        row["*상담시간(시)"] = "0"
        row["*상담시간(분)"] = "40"
        row["*상담사소속"] = "전문상담사"
        row["*상담매체구분"] = "면담"
        row["상담회기"] = "1"
        row[RECORD_ID_COL] = uuid.uuid4().hex
        rows.append(row)
        seq += 1
    
    # 누락된 상담구분 타입이 있으면 추가 (모든 타입 커버)
    remaining_types = set(COUNSELING_TYPES["개인상담"]) - all_types_used
    for ctype in remaining_types:
        student = random.choice(STUDENTS)
        summaries = COUNSELING_SUMMARIES.get(ctype, COUNSELING_SUMMARIES["기타"])
        details = COUNSELING_DETAILS.get(ctype, COUNSELING_DETAILS["기타"])
        date = generate_dates_in_range(f"{YEAR}0301", f"{YEAR}0520", 1)[0]
        
        row = {col: "" for col in SHEET_SCHEMAS["개인상담"]}
        row["순번"] = str(seq)
        row["학번"] = student["sid"]
        row["이름"] = student["name"]
        row["상담시간"] = random.choice(COUNSELING_TIMES)
        row["*상담분류"] = "전문상담"
        row["*Wee클래스"] = "Wee클래스"
        row["*대분류"] = "상담"
        row["*중분류"] = "개인상담"
        row["*상담구분"] = ctype
        row["*상담인원"] = "1"
        row["*학년도"] = date[:4]
        row["*상담일자"] = date
        row["학년"] = f"{student['grade']}학년"
        row["성별"] = student["gender"]
        row["*상담제목"] = random.choice(summaries)
        row["*상담내용"] = row["*상담제목"]
        row["상담내용(상세)"] = random.choice(details)
        row["*상담시간(시)"] = "0"
        row["*상담시간(분)"] = "40"
        row["*상담사소속"] = "전문상담사"
        row["*상담매체구분"] = "면담"
        row["상담회기"] = "1"
        row[RECORD_ID_COL] = uuid.uuid4().hex
        rows.append(row)
        seq += 1
    
    return rows


# ─────────────────────────────────────────────────────
# 8. 집단상담 데이터 생성
# ─────────────────────────────────────────────────────
def generate_group_counseling():
    """
    집단상담 시트 데이터 생성
    - 학번은 콤마 구분으로 여러 학생 포함
    - 단일 학생 레코드도 포함 (테스트 다양성)
    - 프로그램명, 목표, 회기 포함
    """
    rows = []
    seq = 1
    
    # 학년별 그룹 상담 (같은 학년 학생들 묶기)
    grade_groups = {}
    for s in STUDENTS:
        g = s["grade"]
        if g not in grade_groups:
            grade_groups[g] = []
        grade_groups[g].append(s)
    
    for grade, students_in_grade in grade_groups.items():
        if len(students_in_grade) < 3:
            continue
        
        # 각 학년에서 1~2개 집단상담 프로그램
        num_programs = random.randint(1, 2)
        for _ in range(num_programs):
            program = random.choice(GROUP_PROGRAM_NAMES)
            num_sessions = random.randint(2, 5)
            group_size = random.randint(3, min(6, len(students_in_grade)))
            group_members = random.sample(students_in_grade, group_size)
            
            dates = generate_dates_in_range(f"{YEAR}0310", f"{YEAR}0515", num_sessions)
            
            for session_num, date in enumerate(dates, 1):
                sids = ", ".join([m["sid"] for m in group_members])
                genders = set(m["gender"] for m in group_members)
                gender_str = "혼성" if len(genders) > 1 else list(genders)[0]
                counseling_type = random.choice(GROUP_COUNSELING_TYPES)
                
                summaries = COUNSELING_SUMMARIES.get(counseling_type, COUNSELING_SUMMARIES["기타"])
                summary = random.choice(summaries)
                
                row = {col: "" for col in SHEET_SCHEMAS["집단상담(또래상담, 학급별 집단)"]}
                row["순번"] = str(seq)
                row["학번"] = sids
                row["프로그램명"] = program
                row["상담시간"] = "13:10~13:50"
                row["상담회기"] = str(session_num)
                row["목표"] = summary
                row["기타 및 특이사항"] = ""
                row["*상담분류"] = "전문상담"
                row["*Wee클래스"] = "Wee클래스"
                row["*대분류"] = "상담"
                row["*중분류"] = "집단상담"
                row["*상담구분"] = counseling_type
                row["*상담인원"] = str(group_size)
                row["*학년도"] = date[:4]
                row["*상담일자"] = date
                row["학년"] = f"{grade}학년"
                row["성별"] = gender_str
                row["*상담제목"] = summary
                row["*상담내용"] = summary
                row["상담내용(상세)"] = f"{program} {session_num}회기 진행. 참가 학생 {group_size}명. 주제: {summary}"
                row["*상담시간(시)"] = "0"
                row["*상담시간(분)"] = "50"
                row["*상담사소속"] = "전문상담사"
                row["*상담매체구분"] = "면담"
                row[RECORD_ID_COL] = uuid.uuid4().hex
                rows.append(row)
                seq += 1
    
    # 단일 학생 집단상담 (또래상담) - 2~3건
    for _ in range(3):
        student = random.choice(STUDENTS)
        date = generate_dates_in_range(f"{YEAR}0401", f"{YEAR}0510", 1)[0]
        
        row = {col: "" for col in SHEET_SCHEMAS["집단상담(또래상담, 학급별 집단)"]}
        row["순번"] = str(seq)
        row["학번"] = student["sid"]
        row["프로그램명"] = "또래상담"
        row["상담시간"] = "12:30~13:10"
        row["상담회기"] = "1"
        row["목표"] = "또래 관계 향상"
        row["*상담분류"] = "전문상담"
        row["*Wee클래스"] = "Wee클래스"
        row["*대분류"] = "상담"
        row["*중분류"] = "집단상담"
        row["*상담구분"] = "대인관계"
        row["*상담인원"] = "1"
        row["*학년도"] = date[:4]
        row["*상담일자"] = date
        row["학년"] = f"{student['grade']}학년"
        row["성별"] = student["gender"]
        row["*상담제목"] = "또래상담 활동"
        row["*상담내용"] = "또래상담 활동"
        row["상담내용(상세)"] = f"또래상담 프로그램 진행. {student['name']} 학생 참여."
        row["*상담시간(시)"] = "0"
        row["*상담시간(분)"] = "40"
        row["*상담사소속"] = "전문상담사"
        row["*상담매체구분"] = "면담"
        row[RECORD_ID_COL] = uuid.uuid4().hex
        rows.append(row)
        seq += 1
    
    # 오늘 날짜 집단상담 1건 (통계 테스트)
    group = random.sample(STUDENTS[10:20], 4)
    sids = ", ".join([m["sid"] for m in group])
    row = {col: "" for col in SHEET_SCHEMAS["집단상담(또래상담, 학급별 집단)"]}
    row["순번"] = str(seq)
    row["학번"] = sids
    row["프로그램명"] = "자아 존중감 향상 집단상담"
    row["상담시간"] = "14:00~14:40"
    row["상담회기"] = "1"
    row["목표"] = "자기 이해와 존중"
    row["*상담분류"] = "전문상담"
    row["*Wee클래스"] = "Wee클래스"
    row["*대분류"] = "상담"
    row["*중분류"] = "집단상담"
    row["*상담구분"] = "성격"
    row["*상담인원"] = "4"
    row["*학년도"] = YEAR
    row["*상담일자"] = TODAY
    row["학년"] = "혼합"
    row["성별"] = "혼성"
    row["*상담제목"] = "자아 존중감 향상 프로그램 1회기"
    row["*상담내용"] = "자아 존중감 향상 프로그램 1회기"
    row["상담내용(상세)"] = "자아 존중감 향상 집단상담 1회기 진행. 참가 학생 4명. 자기 소개 및 강점 발견 활동."
    row["*상담시간(시)"] = "0"
    row["*상담시간(분)"] = "50"
    row["*상담사소속"] = "전문상담사"
    row["*상담매체구분"] = "면담"
    row[RECORD_ID_COL] = uuid.uuid4().hex
    rows.append(row)
    
    return rows


# ─────────────────────────────────────────────────────
# 9. 보호자상담 데이터 생성
# ─────────────────────────────────────────────────────
def generate_guardian_counseling():
    """
    보호자상담 시트 데이터 생성
    - 10~15명 학생의 보호자 상담
    - 오늘 날짜 포함
    """
    rows = []
    seq = 1
    
    guardian_students = random.sample(STUDENTS, 12)
    
    for student in guardian_students:
        num_sessions = random.choice([1, 1, 1, 2, 2, 3])
        dates = generate_dates_in_range(f"{YEAR}0301", f"{YEAR}0520", num_sessions)
        
        for date in dates:
            counseling_type = random.choice(COUNSELING_TYPES["보호자상담"])
            summaries = COUNSELING_SUMMARIES.get(counseling_type, COUNSELING_SUMMARIES["기타"])
            details = COUNSELING_DETAILS.get(counseling_type, COUNSELING_DETAILS["기타"])
            
            row = {col: "" for col in SHEET_SCHEMAS["보호자상담"]}
            row["순번"] = str(seq)
            row["학번"] = student["sid"]
            row["이름"] = student["name"]
            row["상담시간"] = random.choice(COUNSELING_TIMES)
            row["기타 및 특이사항"] = ""
            row["*상담분류"] = "전문상담"
            row["*Wee클래스"] = "Wee클래스"
            row["*대분류"] = "상담"
            row["*중분류"] = "학부모상담"
            row["*상담구분"] = counseling_type
            row["*상담인원"] = "1"
            row["*학년도"] = date[:4]
            row["*상담일자"] = date
            row["학년"] = f"{student['grade']}학년"
            row["성별"] = student["gender"]
            row["*상담제목"] = random.choice(summaries)
            row["*상담내용"] = row["*상담제목"]
            row["상담내용(상세)"] = random.choice(details)
            row["*상담시간(시)"] = "0"
            row["*상담시간(분)"] = random.choice(["30", "40", "50"])
            row["*상담사소속"] = "전문상담사"
            row["*상담매체구분"] = random.choice(["면담", "전화"])
            row[RECORD_ID_COL] = uuid.uuid4().hex
            rows.append(row)
            seq += 1
    
    # 오늘 날짜 보호자상담 2건
    for student in random.sample(STUDENTS[:20], 2):
        counseling_type = random.choice(COUNSELING_TYPES["보호자상담"])
        summaries = COUNSELING_SUMMARIES.get(counseling_type, COUNSELING_SUMMARIES["기타"])
        details = COUNSELING_DETAILS.get(counseling_type, COUNSELING_DETAILS["기타"])
        
        row = {col: "" for col in SHEET_SCHEMAS["보호자상담"]}
        row["순번"] = str(seq)
        row["학번"] = student["sid"]
        row["이름"] = student["name"]
        row["상담시간"] = "15:30~16:10"
        row["*상담분류"] = "전문상담"
        row["*Wee클래스"] = "Wee클래스"
        row["*대분류"] = "상담"
        row["*중분류"] = "학부모상담"
        row["*상담구분"] = counseling_type
        row["*상담인원"] = "1"
        row["*학년도"] = YEAR
        row["*상담일자"] = TODAY
        row["학년"] = f"{student['grade']}학년"
        row["성별"] = student["gender"]
        row["*상담제목"] = random.choice(summaries)
        row["*상담내용"] = row["*상담제목"]
        row["상담내용(상세)"] = random.choice(details)
        row["*상담시간(시)"] = "0"
        row["*상담시간(분)"] = "40"
        row["*상담사소속"] = "전문상담사"
        row["*상담매체구분"] = "면담"
        row[RECORD_ID_COL] = uuid.uuid4().hex
        rows.append(row)
        seq += 1
    
    return rows


# ─────────────────────────────────────────────────────
# 10. 교원자문 데이터 생성
# ─────────────────────────────────────────────────────
def generate_teacher_consultation():
    """
    교원자문 시트 데이터 생성
    - 5~8명 학생 관련 교원자문
    """
    rows = []
    seq = 1
    
    consultation_students = random.sample(STUDENTS, 7)
    
    for student in consultation_students:
        counseling_type = random.choice(COUNSELING_TYPES["교원자문"])
        summaries = COUNSELING_SUMMARIES.get(counseling_type, COUNSELING_SUMMARIES["기타"])
        details = COUNSELING_DETAILS.get(counseling_type, COUNSELING_DETAILS["기타"])
        date = generate_dates_in_range(f"{YEAR}0301", f"{YEAR}0520", 1)[0]
        
        row = {col: "" for col in SHEET_SCHEMAS["교원자문"]}
        row["순번"] = str(seq)
        row["학번"] = student["sid"]
        row["이름"] = student["name"]
        row["상담시간"] = random.choice(COUNSELING_TIMES)
        row["기타 및 특이사항"] = random.choice(["", "", "담임교사 의견 포함"])
        row["*상담분류"] = "전문상담"
        row["*Wee클래스"] = "Wee클래스"
        row["*대분류"] = "자문"
        row["*중분류"] = "교원자문"
        row["*상담구분"] = counseling_type
        row["*상담인원"] = "1"
        row["*학년도"] = date[:4]
        row["*상담일자"] = date
        row["학년"] = f"{student['grade']}학년"
        row["성별"] = student["gender"]
        row["*상담제목"] = random.choice(summaries)
        row["*상담내용"] = row["*상담제목"]
        row["상담내용(상세)"] = random.choice(details)
        row["*상담시간(시)"] = "0"
        row["*상담시간(분)"] = "30"
        row["*상담사소속"] = "전문상담사"
        row["*상담매체구분"] = "면담"
        row[RECORD_ID_COL] = uuid.uuid4().hex
        rows.append(row)
        seq += 1
    
    return rows


# ─────────────────────────────────────────────────────
# 11. 의뢰 데이터 생성
# ─────────────────────────────────────────────────────
def generate_referrals():
    """
    의뢰 시트 데이터 생성
    - 3~5명 학생 관련 의뢰
    - 오늘 날짜 1건 포함
    """
    rows = []
    seq = 1
    
    referral_students = random.sample(STUDENTS, 4)
    
    for student in referral_students:
        counseling_type = random.choice(COUNSELING_TYPES["의뢰"])
        summaries = COUNSELING_SUMMARIES.get(counseling_type, COUNSELING_SUMMARIES["기타"])
        details = COUNSELING_DETAILS.get(counseling_type, COUNSELING_DETAILS["기타"])
        date = generate_dates_in_range(f"{YEAR}0301", f"{YEAR}0520", 1)[0]
        
        row = {col: "" for col in SHEET_SCHEMAS["의뢰(정서행동의뢰, 자문의 의뢰 등)"]}
        row["순번"] = str(seq)
        row["학번"] = student["sid"]
        row["이름"] = student["name"]
        row["상담시간"] = random.choice(COUNSELING_TIMES)
        row["기타 및 특이사항"] = random.choice(["", "정서행동특성검사 관심군", "보호자 동의 완료"])
        row["*상담분류"] = "전문상담"
        row["*Wee클래스"] = "Wee클래스"
        row["*대분류"] = "의뢰"
        row["*중분류"] = "교내외의뢰"
        row["*상담구분"] = counseling_type
        row["*상담인원"] = "1"
        row["*학년도"] = date[:4]
        row["*상담일자"] = date
        row["학년"] = f"{student['grade']}학년"
        row["성별"] = student["gender"]
        row["*상담제목"] = random.choice(summaries)
        row["*상담내용"] = row["*상담제목"]
        row["상담내용(상세)"] = random.choice(details)
        row["*상담시간(시)"] = "0"
        row["*상담시간(분)"] = "30"
        row["*상담사소속"] = "전문상담사"
        row["*상담매체구분"] = "면담"
        row[RECORD_ID_COL] = uuid.uuid4().hex
        rows.append(row)
        seq += 1
    
    # 오늘 날짜 의뢰 1건
    student = random.choice(STUDENTS)
    row = {col: "" for col in SHEET_SCHEMAS["의뢰(정서행동의뢰, 자문의 의뢰 등)"]}
    row["순번"] = str(seq)
    row["학번"] = student["sid"]
    row["이름"] = student["name"]
    row["상담시간"] = "11:30~12:10"
    row["기타 및 특이사항"] = "긴급 의뢰"
    row["*상담분류"] = "전문상담"
    row["*Wee클래스"] = "Wee클래스"
    row["*대분류"] = "의뢰"
    row["*중분류"] = "교내외의뢰"
    row["*상담구분"] = "외부전문가에게 상담의뢰"
    row["*상담인원"] = "1"
    row["*학년도"] = YEAR
    row["*상담일자"] = TODAY
    row["학년"] = f"{student['grade']}학년"
    row["성별"] = student["gender"]
    row["*상담제목"] = "정서행동 전문기관 의뢰"
    row["*상담내용"] = "정서행동 전문기관 의뢰"
    row["상담내용(상세)"] = "정서행동특성검사 관심군으로 분류된 학생을 외부 전문 상담 기관에 의뢰함. 보호자 동의서 확보 완료."
    row["*상담시간(시)"] = "0"
    row["*상담시간(분)"] = "30"
    row["*상담사소속"] = "전문상담사"
    row["*상담매체구분"] = "면담"
    row[RECORD_ID_COL] = uuid.uuid4().hex
    rows.append(row)
    
    return rows


# ─────────────────────────────────────────────────────
# 12. 엑셀 파일 생성
# ─────────────────────────────────────────────────────
def create_test_excel(output_path):
    """테스트 엑셀 파일 생성"""
    print(f"테스트 엑셀 파일 생성 중... ({output_path})")
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # 스타일 정의
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(bold=True, size=10)
    
    # 시트별 데이터 생성
    sheet_data = {
        "개인상담": generate_individual_counseling(),
        "집단상담(또래상담, 학급별 집단)": generate_group_counseling(),
        "보호자상담": generate_guardian_counseling(),
        "교원자문": generate_teacher_consultation(),
        "의뢰(정서행동의뢰, 자문의 의뢰 등)": generate_referrals(),
    }
    
    for sheet_name in SHEET_SCHEMAS:
        schema = SHEET_SCHEMAS[sheet_name]
        ws = wb.create_sheet(title=sheet_name)
        
        # 헤더 행 작성
        for col_idx, col_name in enumerate(schema, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # 예시 행 추가
        example_row = create_example_row(schema, sheet_name)
        for col_idx, col_name in enumerate(schema, 1):
            cell = ws.cell(row=2, column=col_idx, value=example_row.get(col_name, ""))
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # 데이터 행 추가
        data_rows = sheet_data.get(sheet_name, [])
        for row_idx, row_data in enumerate(data_rows, 3):  # 3행부터 (1: 헤더, 2: 예시)
            for col_idx, col_name in enumerate(schema, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
                cell.alignment = center_alignment
                cell.border = thin_border
        
        # _record_id 컬럼 숨김 처리
        for col_idx, col_name in enumerate(schema, 1):
            if col_name == RECORD_ID_COL:
                ws.column_dimensions[get_column_letter(col_idx)].hidden = True
                break
        
        # 열 너비 설정
        col_widths = {
            "순번": 6, "학번": 10, "이름": 12, "상담시간": 16,
            "기타 및 특이사항": 20, "*상담분류": 12, "*Wee클래스": 12,
            "*대분류": 10, "*중분류": 12, "*상담구분": 25,
            "*상담인원": 10, "*학년도": 10, "*상담일자": 12,
            "학년": 8, "성별": 8, "*상담제목": 30, "*상담내용": 30,
            "상담내용(상세)": 50, "*상담시간(시)": 12, "*상담시간(분)": 12,
            "*상담사소속": 14, "*상담매체구분": 14, "상담회기": 10,
            "프로그램명": 20, "목표": 25, RECORD_ID_COL: 35
        }
        for col_idx, col_name in enumerate(schema, 1):
            width = col_widths.get(col_name, 15)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        print(f"  [{sheet_name}] - 예시 행 1개 + 데이터 {len(data_rows)}행 생성 완료")
    
    # 저장
    wb.save(output_path)
    wb.close()
    print(f"\n✅ 테스트 파일 생성 완료: {output_path}")
    
    # 통계 출력
    total_records = sum(len(v) for v in sheet_data.values())
    unique_students = set()
    for rows in sheet_data.values():
        for row in rows:
            name = row.get("이름", "")
            sid = row.get("학번", "")
            if name and "," not in sid:
                unique_students.add((name, sid))
    
    print(f"\n📊 생성된 테스트 데이터 요약:")
    print(f"   총 레코드 수: {total_records}건")
    print(f"   고유 학생 수: {len(unique_students)}명")
    print(f"   오늘 날짜({TODAY}) 상담 포함: 개인상담 4건, 집단상담 1건, 보호자상담 2건, 의뢰 1건")
    print(f"\n   시트별 레코드 수:")
    for sheet_name, rows in sheet_data.items():
        short = sheet_name[:10] + "..." if len(sheet_name) > 13 else sheet_name
        print(f"     - {short}: {len(rows)}건")
    
    print(f"\n🔍 테스트 가능한 기능:")
    print(f"   ✓ 동명이인 검색 (김민수: 1학년 1101, 2학년 2101)")
    print(f"   ✓ 초성 검색 (ㄱㅁㅅ → 김민수, 구미래)")
    print(f"   ✓ 학번 검색 (1101, 2101, 6304 등)")
    print(f"   ✓ 오늘 통계 (개인 4건 + 집단 1건 + 보호자 2건 + 의뢰 1건)")
    print(f"   ✓ 다회기 상담 (2~8회기 학생 12명)")
    print(f"   ✓ 단회기 상담 (15명)")
    print(f"   ✓ 집단상담 (콤마 구분 학번, 단일/복수 학생)")
    print(f"   ✓ 모든 상담구분 타입 커버 (14종)")
    print(f"   ✓ 다양한 상담매체 (면담/전화/사이버)")
    print(f"   ✓ 학년별 분포 (1~6학년)")
    print(f"   ✓ 성별 분포 (남/여 균등)")
    print(f"   ✓ 드문 성씨 (구미래, 탁현우, 피서진)")
    print(f"   ✓ 특이 이름 (홍길동)")
    print(f"   ✓ 예시 행 포함 (각 시트 첫 행)")
    print(f"   ✓ UUID (_record_id) 자동 생성 및 숨김 처리")


if __name__ == "__main__":
    # data 디렉토리에 테스트 파일 생성
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)
    output_path = os.path.join(project_dir, "data", "테스트_상담일지(50명).xlsx")
    
    random.seed(42)  # 재현 가능성을 위한 시드 고정
    create_test_excel(output_path)
