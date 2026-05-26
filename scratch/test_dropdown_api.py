import os
import sys
import shutil
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

# 로컬 backend 경로 추가
sys.path.append(r"c:\Coding\Projects\School\CounselingLog_Electron")

from backend.utils.validation_parser import parse_validation_options
from backend.core.constants import SHEET_SCHEMAS

def setup_mock_excel(file_path):
    print("Setting up mock excel file with validation types...")
    wb = openpyxl.Workbook()
    ws_main = wb.active
    ws_main.title = "개인상담"
    
    # 1. 개인상담 - 직접 입력 목록 형태
    headers = SHEET_SCHEMAS["개인상담"]
    for col_idx, col_name in enumerate(headers, 1):
        ws_main.cell(row=1, column=col_idx, value=col_name)
    ws_main.cell(row=2, column=headers.index("*상담구분")+1, value="기타")
    
    dv1 = DataValidation(type="list", formula1='"직접입력1,직접입력2,직접입력3"', allow_blank=True)
    wb.active.add_data_validation(dv1)
    dv1.add(ws_main.cell(row=2, column=headers.index("*상담구분")+1))
    
    # 데이터 시트 (Sheet2) 생성하여 참조형 데이터 기입
    ws_data = wb.create_sheet(title="Sheet2")
    # A열: 보호자상담 참조 데이터
    ws_data["A1"] = "참조1"
    ws_data["A2"] = "참조2"
    ws_data["A3"] = "참조3"
    # B열: 교원자문 네임드 레인지 참조 데이터
    ws_data["B1"] = "이름정의1"
    ws_data["B2"] = "이름정의2"
    # C열: 집단상담 전체 열 참조 데이터 (C:C)
    ws_data["C2"] = "전체열1"
    ws_data["C3"] = "전체열2"
    
    # 2. 보호자상담 - 셀 범위 참조 형태
    ws_guardian = wb.create_sheet(title="보호자상담")
    for col_idx, col_name in enumerate(SHEET_SCHEMAS["보호자상담"], 1):
        ws_guardian.cell(row=1, column=col_idx, value=col_name)
    
    dv2 = DataValidation(type="list", formula1="Sheet2!$A$1:$A$3", allow_blank=True)
    ws_guardian.add_data_validation(dv2)
    dv2.add(ws_guardian.cell(row=2, column=SHEET_SCHEMAS["보호자상담"].index("*상담구분")+1))
    
    # 3. 교원자문 - 이름 정의 범위 형태
    ws_teacher = wb.create_sheet(title="교원자문")
    for col_idx, col_name in enumerate(SHEET_SCHEMAS["교원자문"], 1):
        ws_teacher.cell(row=1, column=col_idx, value=col_name)
        
    # 정의된 이름 추가
    new_name = openpyxl.workbook.defined_name.DefinedName("상담옵션이름", attr_text="Sheet2!$B$1:$B$2")
    wb.defined_names["상담옵션이름"] = new_name
    
    dv3 = DataValidation(type="list", formula1="=상담옵션이름", allow_blank=True)
    ws_teacher.add_data_validation(dv3)
    dv3.add(ws_teacher.cell(row=2, column=SHEET_SCHEMAS["교원자문"].index("*상담구분")+1))
    
    # 4. 집단상담 - 전체 열 참조 형태 (C:C)
    ws_group = wb.create_sheet(title="집단상담(또래상담, 학급별 집단)")
    for col_idx, col_name in enumerate(SHEET_SCHEMAS["집단상담(또래상담, 학급별 집단)"], 1):
        ws_group.cell(row=1, column=col_idx, value=col_name)
        
    dv4 = DataValidation(type="list", formula1="Sheet2!$C:$C", allow_blank=True)
    ws_group.add_data_validation(dv4)
    dv4.add(ws_group.cell(row=2, column=SHEET_SCHEMAS["집단상담(또래상담, 학급별 집단)"].index("*상담구분")+1))
    
    # 5. 의뢰시트 - 유효성 검사 미작성 (Fallback 확인용)
    ws_request = wb.create_sheet(title="의뢰(정서행동의뢰, 자문의 의뢰 등)")
    for col_idx, col_name in enumerate(SHEET_SCHEMAS["의뢰(정서행동의뢰, 자문의 의뢰 등)"], 1):
        ws_request.cell(row=1, column=col_idx, value=col_name)
    
    wb.save(file_path)
    wb.close()
    print("Mock excel setup complete.")

def test_validation_options():
    test_excel = r"c:\Coding\Projects\School\CounselingLog_Electron\scratch\test_validation_target.xlsx"
    setup_mock_excel(test_excel)
    
    default_options = {
        "개인상담": ["개인기본"],
        "보호자상담": ["보호자기본"],
        "교원자문": ["교원기본"],
        "의뢰(정서행동의뢰, 자문의 의뢰 등)": ["의뢰기본"]
    }
    
    print("\nRunning parse_validation_options...")
    options_map = parse_validation_options(test_excel, default_options)
    
    # 검증 1: 개인상담 (직접 입력 목록)
    indiv_res = options_map.get("개인상담")
    print(f"개인상담 결과: {indiv_res}")
    assert indiv_res["source"] == "excel"
    assert indiv_res["options"] == ["직접입력1", "직접입력2", "직접입력3"]
    
    # 검증 2: 보호자상담 (범위 참조)
    guardian_res = options_map.get("보호자상담")
    print(f"보호자상담 결과: {guardian_res}")
    assert guardian_res["source"] == "excel"
    assert guardian_res["options"] == ["참조1", "참조2", "참조3"]
    
    # 검증 3: 교원자문 (이름 정의 범위)
    teacher_res = options_map.get("교원자문")
    print(f"교원자문 결과: {teacher_res}")
    assert teacher_res["source"] == "excel"
    assert teacher_res["options"] == ["이름정의1", "이름정의2"]
    
    # 검증 4: 집단상담 (전체 열 참조 C:C)
    group_res = options_map.get("집단상담(또래상담, 학급별 집단)")
    print(f"집단상담 결과: {group_res}")
    assert group_res["source"] == "excel"
    # C:C 열에서 값이 있는 셀만 중복 없이 추출
    assert "전체열1" in group_res["options"]
    assert "전체열2" in group_res["options"]
    
    # 검증 5: 의뢰 (Fallback)
    request_res = options_map.get("의뢰(정서행동의뢰, 자문의 의뢰 등)")
    print(f"의뢰 결과: {request_res}")
    assert request_res["source"] == "fallback"
    assert request_res["options"] == ["의뢰기본"]
    
    if os.path.exists(test_excel):
        os.remove(test_excel)
    print("\nALL EXCEL DROPDOWN PARSING TESTS PASSED SUCCESSFULLY!")
    return True

if __name__ == "__main__":
    try:
        if test_validation_options():
            sys.exit(0)
    except Exception as e:
        print(f"TEST FAILED: {e}")
        sys.exit(1)
