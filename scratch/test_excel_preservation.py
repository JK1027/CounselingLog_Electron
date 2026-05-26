import os
import sys
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# 로컬 backend 경로 추가
sys.path.append(r"c:\Coding\Projects\School\CounselingLog_Electron")

from backend.repositories.excel_repository import ExcelRepository

def setup_mock_excel(file_path):
    print("Setting up mock excel file with formats...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "개인상담"
    
    headers = ["순번", "이름", "학번", "학년", "성별", "*상담일자", "*상담제목", "상담내용(상세)", "_record_id"]
    for col_idx, col_name in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # 1. 예시 행 (2행) 추가
    row2 = ["예시", "예시학생", "00000", "1학년", "남", "2026-05-26", "예시제목", "예시상세", "uuid_1"]
    for col_idx, val in enumerate(row2, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = Font(name="Malgun Gothic", size=11, bold=True, color="FF0000")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.row_dimensions[2].height = 25
    
    # 2. 실제 학생 행 (3행) 추가
    row3 = ["1", "홍길동", "10101", "1학년", "남", "2026-05-26", "상담제목1", "상담상세1", "uuid_2"]
    for col_idx, val in enumerate(row3, 1):
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.font = Font(name="Malgun Gothic", size=11, bold=True, color="FF0000")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.row_dimensions[3].height = 25
    
    # 자동 필터 설정 (A1:I3)
    ws.auto_filter.ref = "A1:I3"
    
    # 데이터 유효성 검사 (드롭박스) 설정 (E2:E3)
    dv = DataValidation(type="list", formula1='"남,여"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["E2"])
    dv.add(ws["E3"])
    
    for sheet_name in ["집단상담(또래상담, 학급별 집단)", "보호자상담", "교원자문", "의뢰(정서행동의뢰, 자문의 의뢰 등)"]:
        wb.create_sheet(title=sheet_name)
        
    wb.save(file_path)
    wb.close()
    print("Mock excel setup complete.")

def test_preservation():
    test_excel = r"c:\Coding\Projects\School\CounselingLog_Electron\scratch\test_preservation_target.xlsx"
    setup_mock_excel(test_excel)
    
    repo = ExcelRepository()
    repo.load_data(test_excel)
    
    # 1. 학생 정보 수정 테스트 (save_all_data_to_excel 유발)
    print("\n[Test 1] Modifying student info (update_student_info)...")
    success, err = repo.update_student_info(
        old_name="홍길동",
        old_student_id="10101",
        new_name="김철수",
        new_student_id="10102",
        grade="2",
        gender="여"
    )
    if not success:
        print(f"Update failed: {err}")
        return False
    print("Update successful!")
    
    # 검증 1
    wb = openpyxl.load_workbook(test_excel)
    ws = wb["개인상담"]
    
    print("Checking Test 1 formatting retention on Row 3:")
    cell_e3 = ws["E3"]
    print(f"  E3 Value: {cell_e3.value} (Expected: 여)")
    print(f"  E3 Bold font: {cell_e3.font.bold} (Expected: True)")
    print(f"  E3 Font color: {cell_e3.font.color.rgb} (Expected: FFFF0000 or similar)")
    print(f"  E3 Fill color: {cell_e3.fill.start_color.rgb} (Expected: FFFFFFF00 or FFFF00)")
    print(f"  E3 Row Height (Row 3): {ws.row_dimensions[3].height} (Expected: 25)")
    print(f"  AutoFilter range: {ws.auto_filter.ref} (Expected: A1:I3)")
    print(f"  Data validations count: {len(ws.data_validations.dataValidation)}")
    
    found_dv = False
    for dv in ws.data_validations.dataValidation:
        print(f"  DV range: {dv.sqref}")
        if "E3" in str(dv.sqref):
            print(f"  Found E3 Validation! Formula: {dv.formula1}")
            found_dv = True
    
    if not found_dv:
        print("  ERROR: Data validation on E3 was lost!")
        return False
        
    wb.close()
    
    # 2. 신규 행 추가 테스트 (append_new_row_to_excel 유발)
    print("\n[Test 2] Appending new row (append_new_row_to_excel)...")
    import pandas as pd
    new_row = pd.DataFrame([{
        "순번": "2",
        "이름": "이영희",
        "학번": "20202",
        "학년": "2학년",
        "성별": "여",
        "*상담일자": "2026-05-26",
        "*상담제목": "새상담",
        "상담내용(상세)": "새상세",
        "_record_id": "uuid_3"
    }])
    
    success, err = repo.append_new_row_to_excel("개인상담", new_row)
    if not success:
        print(f"Append failed: {err}")
        return False
    print("Append successful!")
    
    # 검증 2
    wb = openpyxl.load_workbook(test_excel)
    ws = wb["개인상담"]
    
    print("Checking Test 2 formatting and range expansion on Row 4:")
    cell_e4 = ws["E4"]
    print(f"  E4 Value: {cell_e4.value} (Expected: 여)")
    print(f"  E4 Bold font: {cell_e4.font.bold} (Expected: True - cloned from E3)")
    print(f"  E4 Fill color: {cell_e4.fill.start_color.rgb} (Expected: FFFFFFF00 - cloned from E3)")
    print(f"  E4 Row Height (Row 4): {ws.row_dimensions[4].height} (Expected: 25 - cloned from Row 3)")
    print(f"  AutoFilter range: {ws.auto_filter.ref} (Expected: A1:I4 - expanded to Row 4)")
    print(f"  Data validations count: {len(ws.data_validations.dataValidation)}")
    
    found_e4_dv = False
    for dv in ws.data_validations.dataValidation:
        print(f"  DV sqref: {dv.sqref}")
        if "E4" in str(dv.sqref):
            print(f"  Found E4 in Data Validation range! Formula: {dv.formula1}")
            found_e4_dv = True
            
    if not found_e4_dv:
        print("  ERROR: Data validation did not expand to E4!")
        return False
        
    wb.close()
    
    # Cleanup
    if os.path.exists(test_excel):
        os.remove(test_excel)
    print("\nALL PRESERVATION TESTS PASSED SUCCESSFULLY!")
    return True

if __name__ == "__main__":
    if test_preservation():
        sys.exit(0)
    else:
        sys.exit(1)
